#!/usr/bin/env python2
# -*- coding: utf-8 -*-
'''
This task was made for psychopy 1.85.4
Shows the videos randomly (different order for each participant)
The task saves an Excel file with reaction times inside the folder 'data'
press esc to exit the task
Author: Alberto Gonzalez (albertojac.gonzalez@gmail.com)
'''
from __future__ import division
import os, csv, random
from psychopy import visual, core, event, data, gui
from random import shuffle
from openpyxl.workbook import Workbook

# task instructions
instructionsText1 = 'Press Q for Yes and M for No\n\n      Press space to start'
instructionsText2 = 'Press Q for No and M for Yes\n\n      Press space to start'

FULLSCR = True
ISI = 3



# initial gui
expInfo = {'1 SubjectNumber':'', '2 Group': ['Control', 'Test']}
expName = 'GesturesTask'
infoDlg = gui.DlgFromDict(dictionary=expInfo, title= expName)
if not infoDlg.OK:
    print 'User Cancelled'

# create a folder to save the data
expInfo['Date'] = data.getDateStr()
if not os.path.isdir('Data'): 
    os.makedirs('Data')  
    
# define the filename
filename = 'Data' + os.path.sep + '%s_%s_%s_Part2' %(expInfo['1 SubjectNumber'], expInfo['2 Group'],expInfo['Date'])

wb = Workbook()
ws = wb.active # active worksheet 
ws.title = str(expInfo['1 SubjectNumber']) + expInfo['2 Group']

ws['A1'] = 'TrialNumber'
ws['B1'] = 'StimulusType'
ws['C1'] = 'RT'
ws['D1'] = 'ResponseKey'
ws['E1'] = 'Response'
ws['F1'] = 'MovieOnset'
ws['G1'] = 'MovieOffset'

# counterbalancing among participants
if int(expInfo['1 SubjectNumber'])%2 == 0:
    instructionsText = instructionsText1
    group = 0
else:
    instructionsText = instructionsText2
    group = 1
    
# load optSeq file
#times = []
#stimType = []
#stimDurat = []
#stimTypeName = []
#with open('OptSeq' + os.path.sep + 'taskOptseq-001.par', 'rb') as f:
#    reader = csv.reader(f, delimiter='\t')
#    for row in reader:
#        times.append(float(row[0][0:9]))
#        stimType.append(int(row[0][11:14]))
#        stimDurat.append(float(row[0][16:24]))
#        stimTypeName.append(row[0][34:].replace(" ", ""))


# video files
IntransClose = ['sMF04v_inM-Mov.m4v',
                         'sMF11v_inM-Mov.m4v',
                         'sMF17v_inM-Mov.m4v',
                         'sMF08v_inM-Mov.m4v',
                         'dMF01v_inM-Mov.mp4',
                         'dMF18v_inM-Mov.m4v']
# shuffle(IntransClose)
IntransFar = ['dMF33l_inI-Mov.m4v',
                     'dMF42l_inI-Mov.m4v',
                     'sMF33l_inl-Mov.m4v',
                     'dMF44l_inI-Mov.m4v',
                     'dMF38l_inI-Mov.m4v',
                     'dMF39l_inI-Mov.m4v']
#shuffle(IntransFar)
PantoClose = ['dMF65v_pa-Mov.m4v',
                     'dMF79v_pa-Mov.m4v',
                     'dMF70v_pa-Mov.m4v',
                     'sMF72v_pa-Mov.m4v',
                     'dMF69v_pa-Mov.m4v',
                     'sMF77v_pa-Mov.m4v',
                     'dMF75v_pa-Mov.m4v',
                     'sMF59v_pa-Mov.m4v',
                     'dMF60v_pa-Mov.m4v',
                     'dMF81v_pa-Mov.m4v',
                     'dMF63v_pa-Mov.m4v',
                     'sMF61v_pa-Mov.m4v',
                     'dMF76v_pa-Mov.m4v',
                     'sMF68v_pa-Mov.m4v',
                     'sMF71v_pa-Mov.m4v',
                     'dMF73v_pa-Mov.m4v',
                     'sMF78v_pa-Mov.m4v',
                     'dMF66v_pa-Mov.m4v']
#shuffle(PantoClose)
PantoFar = ['dMF89l_pa-Mov.m4v',
                     'sMF51l_pa-Mov.m4v',
                     'dMF53l_pa-Mov.m4v',
                     'sMF48l_pa-Mov.m4v',
                     'sMF87l_pa-Mov.m4v',
                     'dMF56l_pa-Mov.m4v',
                     'sMF54l_pa-Mov.m4v',
                     'sMF91l_pa-Mov.m4v',
                     'sMF57l_pa-Mov.m4v',
                     'dMF55l_pa-Mov.m4v',
                     'sMF84l_pa-Mov.m4v',
                     'dMF50l_pa-Mov.m4v',
                     'sMF52l_pa-Mov.m4v',
                     'sMF88l_pa-Mov.m4v']
#shuffle(PantoFar)
MLClose = ['dML11v-Mov.mp4',
                     'sML03v-Mov.m4v',
                     'dML07v-Mov.mp4',
                     'dML06v-Mov.mp4',
                     'sML15v-Mov.m4v',
                     'sML26v-Mov.m4v',
                     'sML19v-Mov.m4v',
                     'sML18v-Mov.m4v',
                     'sML65v-Mov.m4v',
                     'sML13v-Mov.m4v',
                     'dML23v-Mov.mp4',
                     'dML22v-Mov.mp4',
                     'dML29v-Mov.mp4']
#shuffle(MLClose)
MLFar = ['dML57l-Mov.m4v',
                 'dML31l-Mov.mp4',
                 'sML54l-Mov.m4v',
                 'dML37l-Mov.m4v',
                 'dML36l-Mov.m4v',
                 'sML61l-Mov.m4v',
                 'dML41l-Mov.m4v',
                 'sML58l-Mov.m4v',
                 'sML45l-Mov.m4v',
                 'dML50l-Mov.m4v',
                 'dML51l-Mov.m4v']
#shuffle(MLFar)
movies = MLFar + MLClose + PantoFar + PantoClose + IntransFar + IntransClose
shuffle(movies)

win = visual.Window((800, 800), fullscr = FULLSCR)
fixation = visual.ShapeStim(win,lineWidth=6,vertices=((-0.05, 0), (0.05, 0), (0,0), (0,0.08), (0,-0.08)), interpolate = False, closeShape=False)
mov = visual.MovieStim3(win,'movies' + os.path.sep + 'dMF17v_inM-Mov.m4v')
trialClock = core.Clock()
globalClock = core.Clock()

instruct = visual.TextStim(win=win, name='text',text= instructionsText,font='Arial', height=0.12, wrapWidth=1.8, color='white',alignVert='center')

# functions for responses
def clear_all_cue():
    while True:
        if event.getKeys(): event.clearEvents('keyboard')
        else: break

def get_response():
    key_out=[]
    resp=event.getKeys()
    if resp:
        key=resp[0]
        if key=='space':key_out=98
        elif key=='q':key_out= key
        elif key=='m':key_out= key
        elif key=="escape":key_out=99
        return key_out

event.Mouse(visible = False)
instruct.draw()
win.flip()
clear_all_cue()
while True:
    resp = get_response()
    if resp:
        if resp == 98:
            globalClock.reset()
            break
win.flip()
totalTime = 0
# begin the loop
for trial, movie in enumerate(movies):
#    if stimType[trial] == 0: # if null 
#        pass
#    else:
#        if stimType[trial] == 1: # MLClose
#            movFile = MLClose.pop()
#        elif stimType[trial] == 2: # MLFar
#            movFile = MLFar.pop()
#        elif stimType[trial] == 3: # MLFar
#            movFile = PantoClose.pop()
#        elif stimType[trial] == 4: # MLFar
#            movFile = PantoFar.pop()
#        elif stimType[trial] == 5: # MLFar
#            movFile = IntransClose.pop()
#        elif stimType[trial] == 6: # MLFar
#            movFile = IntransFar.pop()

    mov.setMovie('movies' + os.path.sep + movie)
    firstTrial = True
    responded = False
    while mov.status != visual.FINISHED:
        mov.draw()
        win.flip()
        if firstTrial:
            trialClock.reset()
            mov_onset = globalClock.getTime()
            clear_all_cue()
            firstTrial = False
        if not responded:
            resp = get_response()
            if resp:
                if resp == 'm' or resp == 'q':
                    rt = trialClock.getTime()
                    responded = True
                elif resp == 99:
                    wb.save(filename + '_inc.xlsx')
                    win.close()
                    core.quit()

    fixation.draw()
    win.flip()
    mov_offset = globalClock.getTime()
    totalTime = totalTime + mov.duration + ISI
    while globalClock.getTime() < totalTime:
        if not responded:
                resp = get_response()
                if resp:
                    if resp == 'm' or resp == 'q':
                        rt = trialClock.getTime()
                        responded = True
                    elif resp == 99:
                        wb.save(filename + '_inc.xlsx')
                        win.close()
                        core.quit()
        core.wait(.002)
    
    ws['A'+str(trial+2)] = trial+1
    ws['B'+str(trial+2)] = movie
    if responded:
        ws['C'+str(trial+2)] = rt
        ws['D'+str(trial+2)] = resp
        if group == 0 and resp == 'q':
            ws['E'+str(trial+2)] = 'yes'
        elif group == 0 and resp == 'm':
            ws['E'+str(trial+2)] = 'no'
        elif group == 1 and resp == 'q':
            ws['E'+str(trial+2)] = 'no'
        elif group == 1 and resp == 'm':
            ws['E'+str(trial+2)] = 'yes'

    ws['F'+str(trial+2)] = mov_onset
    ws['G'+str(trial+2)] = mov_offset

wb.save(filename + '.xlsx')
win.close()
core.quit()
