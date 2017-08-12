# -*- coding: utf-8 -*-

from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
import os, sys, socket, pyxid
from psychopy import visual, core, data, event, gui, parallel
from numpy.random import random
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook


NTRIALS = 240

# Turn fullscreen off for testing
FULL_SCREEN = True

# task parameters
STIM_TIME = .9
STIM_ISI_SECONDS = 1.7 # + 0.0-0.5
RANDOM_TIME= .5
NSTIM_BLOCK = 40
NBLOCK = 20
BLOCKS = int(NTRIALS/NSTIM_BLOCK)


# Value returned by the Lumina when each of the buttons is pressed
BUTTON_1 = 4
BUTTON_2 = 5
BUTTON_3 = 6


# Task instructions:
task_instructions = u"\
                    Cada pocos segundos, los números (1, 2, 3, o 0) \n\
                    aparecerán en el centro de la pantalla.\n\
                    Un numero siempre será diferente de los otros 2.\n\
                    Presiona la tecla de acuerdo con la identidad,\n\
                    y no la posición, del número diferente.\n\
                    Los valores que corresponden a las teclas son:\n\
                    dedo índice = 1, corazón = 2, anular = 3\n\
                    Responde lo más correctamente y rápido que puedas."




expName = 'MSIT'  # from the Builder filename that created this script

# gui dialogue to get participant id and group
expInfo = {'ID Participante':'',\
           'Grupo': ['FM', 'GC']}

dlg = gui.DlgFromDict(dictionary=expInfo, title=expName, fixed=[])

# if user pressed cancel, quit
if dlg.OK == False:
    core.quit()

expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName

stim_list = ['100', '020', '233', '100', '020', '232', '003', '100', '233',\
            '100', '311', '003', '233', '313', '020', '212', '211', '100', \
            '112', '331', '100', '322', '212', '020', '331', '003', '322', \
            '212','020', '232', '003', '100', '020', '020',  '212', '100', \
            '311', '131', '020', '003', '212', '003', '020', '322', '020', \
            '020', '311', '331', '322', '332', '003', '100', '020', '211', \
            '100', '003', '003', '100', '212', '100', '020', '003', '003', \
            '003', '311', '003', '020', '331', '313', '221', '003', '232', \
            '020', '020', '003', '131', '331', '211', '003', '100', '332', \
            '020', '020', '212', '331', '211', '112', '112', '100', '003', \
            '020', '313', '112', '131', '332', '100', '003', '100', '313', \
            '003', '131', '020', '100', '100', '322', '100', '020', '020', \
            '003', '313', '003', '131', '100', '003', '100', '211', '100', \
            '100', '020', '003', '232', '003', '331', '003', '020', '331', \
            '020', '313', '100', '322', '003', '003', '221', '221', '020', \
            '020', '100', '020', '332', '003', '311', '020', '003', '211', \
            '211', '100', '331', '221', '331', '331', '003', '313', '020', \
            '233', '131', '020', '100', '100', '020', '233', '020', '211', \
            '100', '221', '131', '100', '003', '100', '100', '232', '020', \
            '131', '003', '313', '100', '332', '020', '322', '322', '112', \
            '211', '211', '332', '232', '100', '211', '100', '100', '100', \
            '232', '311', '322', '311', '020', '212', '003', '221', '003', \
            '100', '233', '212', '332', '020', '322', '233', '003', '221', \
            '212', '003', '212', '131', '020', '232', '232', '003', '232', \
            '311', '322', '211', '212', '211', '212', '322', '331', '020', \
            '112', '003', '322', '112', '131', '112', '232', '020', '322', \
            '100', '311', '112', '211', '313', '211', '311', '003', '020', \
            '112', '311', '131', '020', '100', '322', '100', '322', '221', \
            '020', '020', '003', '212', '003', '233', '322', '020', '003', \
            '332', '233', '003', '131', '003', '331', '311', '020', '100', \
            '221', '020', '211', '100', '322', '233', '003', '221', '100', \
            '331', '003', '232', '003', '100', '100', '020', '322', '232', \
            '003', '100', '020', '322', '100', '233', '221', '313', '020', \
            '100', '100', '003', '003', '100', '020', '232', '100', '003', \
            '100', '313', '322', '003', '332', '003', '322', '100', '212', \
            '100', '212', '313', '020', '112', '020', '100', '100', '332', \
            '003', '221', '100', '112', '020', '232', '311', '020', '131', \
            '112', '131', '221', '003', '100', '313', '331', '221', '003', \
            '020', '003', '311', '100', '003', '311', '003', '020', '221', \
            '003', '232', '331', '112', '221', '003', '020', '313', '003', \
            '332', '233', '100', '332', '020', '003', '233', '311', '100', \
            '332', '020', '233', '233', '131', '100', '020', '322', '003', \
            '003', '313', '020', '332', '020', '112', '100', '332', '020', \
            '322', '100', '020', '233', '003', '003', '112', '332', '003', \
            '131', '313', '020', '100']

# keep only the desired number of trials
stim_list = stim_list[0:NTRIALS]



# Define some functions

# function to send triggers by parallel or by serial port if neuroelectrics
# startstim device is avaliable
def send_trigger(code,parallel_port,parallel,s):
    if parallel_port == 1:
        parallel.setData(code)
    if parallel_port == 2 and code:
        trigg = '<TRIGGER>' + str(code)+ '</TRIGGER>'
        s.send(trigg)
    else:
        pass

# clears all cue of responses
def clear_all_cue(dev,cedrus):
    if cedrus:
        while True:
            dev.poll_for_response()
            if dev.response_queue_size():
                dev.clear_response_queue()
            else: break
    else:
        while True:
            if event.getKeys(): event.clearEvents('keyboard')
            else: break


# get responses from keyboard or cedrus if avaliable
def get_response(cedrus,dev,button1,button2,button3):
    key_out = []
    if cedrus:
        dev.poll_for_response()
        if dev.response_queue_size()>0:
            resp = dev.get_next_response()
            if resp['pressed'] == True:
                key = resp['key']
                if key == button1:#codigo que deberia mandar o cedrus lumina400pair ao pulsar o primeiro botón da man dereita
                    key_out = 1
                elif key == button2:#codigo que deberia mandar o cedrus lumina400pair ao pulsar o segundo botón da man esquerda
                    key_out = 2
                elif key == button3:#codigo que deberia mandar o cedrus lumina400pair ao pulsar o primeiro botón da man dereita
                    key_out = 3
                return key_out
        else:
            resp = event.getKeys()
            if resp:
                if resp[0] == 'escape': key_out = 99
                return key_out
    else:
        resp = event.getKeys()
        if resp:
            key = resp[0]
            if key == '1' or key == 'num_1' :key_out = 1#codigo que deberia mandar o cedrus lumina400pair ao pulsar o segundo botón da man esquerda
            elif key == '2' or key == 'num_2' :key_out = 2 #codigo que deberia mandar o cedrus lumina400pair ao pulsar o primeiro botón da man dereita
            elif key == '3' or key == 'num_3' :key_out = 3
            elif key == "escape": key_out = 99
            return key_out

# parallel port
try:
    parallel.setPortAddress(0x378)
    parallel.setData(0)
    parallel_port = 1
except:
    parallel_port = 0
    print "Oops!  No parallel port found"

s = []
if parallel_port == 0:
    try:
        s = socket.socket( socket.AF_INET, socket.SOCK_STREAM )
        s.connect(( "127.0.0.1", 1234 ))
        parallel_port = 2
    except:
        print "No TCP port found"


## initialize communication with the Lumina
try:
    devices = pyxid.get_xid_devices()
    dev = devices[0] # a caixa de respostas sera "dev"
    if dev.is_response_device():
        dev.reset_base_timer()
        dev.reset_rt_timer()
    dev.clear_response_queue()
    cedrus = 1
    print "Cedrus response box found! Maybe you need to change the response keys (see line 121)"

except:
    print "No cedrus response box found, we will use the keyboard"
    cedrus = 0
    dev = None

# Setup files for saving
if not os.path.isdir('MSITData'):
    # if this fails (e.g. permissions) we will get error
    os.makedirs('MSITData' )

filename = 'MSITData' \
    + os.path.sep + '%s_%s_%s' %(expInfo['ID Participante'],expInfo['Grupo'],expInfo['date'])


#########
# EXCEL #
#########

xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]
xlSheet.title = 'MSIT' + str(expInfo['ID Participante']) +str(expInfo['Grupo'])+ str(expInfo['date'])

xlSheet.cell('A1').value = 'stim'
xlSheet.cell('B1').value = '0 = Control; 1 = Interf'
xlSheet.cell('C1').value = 'Stim Duration'
xlSheet.cell('D1').value = '1=Index; 2=middle; 3=ring'
xlSheet.cell('E1').value = 'RT'
xlSheet.cell('F1').value = '0=Incorrect; 1=Correct'
xlSheet.cell('G1').value = 'tBloque'
xlSheet.cell('H1').value = 'tTotal'



# Setup the Window
win = visual.Window(size=(1000, 600), fullscr=FULL_SCREEN, screen=0,
                    allowGUI=False, allowStencil=False, monitor='testMonitor',
                    color='black', colorSpace='rgb')

# Initialize components for instructions
instruct_text = visual.TextStim(win=win, ori=0, text=task_instructions,
                font='Arial',alignHoriz='center', alignVert='center',
                pos=[-0.1, 0], height=0.09, wrapWidth=1.8)
                
start_text = visual.TextStim(win=win, text='Comenzamos. Pulsa\npara continuar',
            font='Arial',  height=0.15,color='white', colorSpace='rgb')
            
rest_text = visual.TextStim(win=win, text='Descanso. Pulsa\npara continuar',
            font='Arial',  height=0.15,color='white', colorSpace='rgb')

practice_text = visual.TextStim(win=win, text='Practica',
            font='Arial',  height=0.15,color='white', colorSpace='rgb')

# fixation cross
fixation = visual.ShapeStim(win,units='cm', lineColor='white', lineWidth=1.2,
            vertices=((-0.5, 0), (0.5, 0), (0,0), (0,0.5), (0,-0.5)),
            interpolate=True, closeShape=False, pos=(0,0))

stim_text = visual.TextStim(win=win, text=' ',
            font='Arial',  height=0.25,color='white', colorSpace='rgb')



########################
# Start the experiment #
########################

# Initialize timers
globalClock = core.Clock()  # to track the time since experiment started
globalClock2 = core.Clock()
trialClock = core.Clock()

instruct_text.draw()
win.flip()
core.wait(1)

while True:
    if get_response(cedrus, dev, BUTTON_1, BUTTON_2, BUTTON_3): #se dou resposta
        break

send_trigger(0, parallel_port, parallel, s)

TRIAL_STIMS = ['100', '020', '003', '131','020','221','233','100','322']
trial_time = 1

practice_text.draw()
win.flip()
core.wait(2)
win.flip()
core.wait(1)
globalClock.reset()
for trial_stim in TRIAL_STIMS:
    stim_text.setText(trial_stim)
    stim_text.setAutoDraw(True)
    while globalClock.getTime() < trial_time - (.005):
        pass
    win.flip()
    trial_time = trial_time+STIM_TIME
    while globalClock.getTime()<trial_time-(.005):
        resp = get_response(cedrus, dev, BUTTON_1, BUTTON_2, BUTTON_3)
        if resp:
            if resp==99:
                win.close()
                core.quit()
    stim_text.setAutoDraw(False)
    win.flip()
    trial_time = trial_time + STIM_ISI_SECONDS + (random()*RANDOM_TIME)

start_text.draw()
win.flip()
core.wait(2)
clear_all_cue(dev,cedrus)

while True:
    if get_response(cedrus, dev, BUTTON_1, BUTTON_2, BUTTON_3): #se dou resposta
        break
win.flip()
clear_all_cue(dev,cedrus)
nEnsaio = 0
core.wait(1)

globalClock2.reset()
for block in range(BLOCKS):
    stim_list2 = stim_list[40*block:40*block+40]
    win.flip()
    globalClock.reset()
    time = 1
    for stim in stim_list2:
        stim_type = []
        stim_text.setText(stim)
        stim_text.draw()
        #fixation.setAutoDraw(False)
        if stim == '100':
            code = 11
            stim_type = 0
        elif stim == '020':
            code = 12
            stim_type = 0
        elif stim == '003':
            code = 13
            stim_type = 0
        elif stim == '221' or stim =='212' or stim =='331' or stim =='313':
            code = 21
            stim_type = 1
        elif stim == '112' or stim =='211' or stim =='332' or stim =='233':
            code = 22
            stim_type = 1
        elif stim == '131' or stim =='311' or stim =='232' or stim =='322':
            code = 23
            stim_type = 1

        #print (str(time) + ' Tempo real:' + str(globalClock.getTime()))
        while globalClock.getTime() < time-(.005):
            pass
        win.flip()
        send_trigger(code,parallel_port,parallel,s)
        trialClock.reset()
        time = time+STIM_TIME
        rt = None
        correcto = 0
        nEnsaio += 1
        while globalClock.getTime() < time-(.005): # presentación de stim
            if trialClock.getTime() < .03:
                clear_all_cue(dev,cedrus)
            if trialClock.getTime() > .03 and not rt:
                resp = get_response(cedrus, dev, BUTTON_1, BUTTON_2, BUTTON_3)
                if resp:
                    rt = trialClock.getTime()
                    send_trigger(resp, parallel_port, parallel, s) # en realidad hai que mandar resp-algo
                    if resp == 99:
                        xlsxWriter.save(filename = filename+'_inc.xlsx')#se presionou escape: salir e gardar o archivo incompleto
                        win.close()
                        core.quit()

        win.flip()
        t_stim = trialClock.getTime()
        time = time + STIM_ISI_SECONDS + (random() * RANDOM_TIME) # 1500 + ms random
        while globalClock.getTime() < time-(.025):
            if not rt:
                resp = get_response(cedrus,dev,BUTTON_1,BUTTON_2,BUTTON_3)
                if resp:
                    rt = trialClock.getTime()
                    send_trigger(resp, parallel_port, parallel,s) # en realidad hai que mandar resp-algo
                    if resp == 99:
                        xlsxWriter.save(filename = filename + '_inc.xlsx')#se presionou escape: salir e gardar o archivo incompleto
                        win.close()
                        core.quit()

        send_trigger(0, parallel_port, parallel, s)
        corr_resp = None
        if rt:
            if cedrus:
                corr_resp=resp 
                if code-20 == resp or code-10 == resp:
                    correcto = 1
            else:
                corr_resp=resp
                if code-20 == resp or code-10 == resp:
                    correcto = 1

        xlSheet.cell('A'+str(nEnsaio+1)).value = stim
        xlSheet.cell('B'+str(nEnsaio+1)).value = stim_type
        xlSheet.cell('C'+str(nEnsaio+1)).value = t_stim
        xlSheet.cell('D'+str(nEnsaio+1)).value = corr_resp
        xlSheet.cell('E'+str(nEnsaio+1)).value = rt
        xlSheet.cell('F'+str(nEnsaio+1)).value = correcto
        xlSheet.cell('G'+str(nEnsaio+1)).value = globalClock.getTime()
        xlSheet.cell('H'+str(nEnsaio+1)).value = globalClock2.getTime()


    # Rest every 2 blocks
    if block<5:
        rest_text.setAutoDraw(True)
        win.flip()
        core.wait(1)
        while True:
            if get_response(cedrus, dev, BUTTON_1, BUTTON_2, BUTTON_3): #se dou resposta
                rest_text.setAutoDraw(False)
                break

rest_text.setText('Gracias')
rest_text.draw()
win.flip()
xlsxWriter.save(filename = filename +'.xlsx')#
core.wait(2)
win.close()
core.quit()
