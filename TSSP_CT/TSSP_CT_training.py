#!/usr/bin/env python2
# -*- coding: utf-8 -*-

from __future__ import division
import serial,os, csv, random
from psychopy import visual, core, event, data, gui, parallel
from openpyxl import Workbook


temp = 46.0 # temperature (a decimal number e.g.: 47.0)
vibration = 20 # intensity of the vibration (an integer value from 0 to 120)


try:
    ser = serial.Serial('/dev/cu.usbmodem11201', 115200)
except:
    ser=False
    print("Oops! No serial port avaliable!!")

expName = 'TSSP_CT_task'

vas_serial = 'v' # trigger VAS1 e VAS12
ts_serial = 's' # trigger bloque TSSP alone
ct_serial= 'w' # trigger bloque CT stroking + TSSP
vi_serial = 'z' # trigger bloque Vibration + TSSP
trial_dur = 3 # duracion de cada trial, 3 segundos
temp_fs = .5 # sample temp serial port each 500ms
ts_trig = 20 # trigger sent by parallel for marking heat in TSSP alone condition
vi_trig = 40 # trigger sent for TSSP+vibration
ct_trig = 60 # trigger sent for TSSP+CT stroking
FULLSCR = True


# define some functions
def send_trigger(code,parallel_port,parallel):
    if parallel_port == 1:
        parallel.setData(code)
        core.wait(0.01)
        parallel.setData(0)
    else:
        pass

def clear_all_cue():
    while True:
        if event.getKeys(): event.clearEvents()
        else: break


def send_data(data_out): # send string by serial
    if ser!=False:
        for i in data_out:
            ser.write(i.encode())
            core.wait(0.002)

def read_data(): # read serial
    if ser!=False:
        incoming=''
        while ser.inWaiting():
            incoming = incoming + ser.read().decode()
        return(incoming)

all_temps = [] # save temperatures and time in this variable
def read_temp():
    if ser!=False:
        incoming=''
        if ser.inWaiting():
            while ser.inWaiting():
                incoming = incoming + ser.read().decode()
            all_temps.append([incoming,globalClock.getTime()])

# Setup parallel port
try:
    parallel.setPortAddress(0x378)#direccion do porto paralelo
    parallel.setData(0)
    parallel_port=1
except:
    parallel_port=0
    print ("Oops!  No parallel port found")

def get_response():
    key_out=[]
    resp = event.getKeys()
    mouse1, mouse2, mouse3 = myMouse.getPressed()
    if resp:
        key=resp[0]
        if key=="escape":
            key_out=99
        return key_out
    if (mouse1):
        key_out=1
        return key_out

# EXCEL
def saveExcel(blockTypes,vas1,vas2,vas3,rts1,rts2,rts3,vas1t,vas2t,vas3t,handsused,suffix):
    dest_filename = 'DataTSSP' + os.path.sep + '%s_%s_%s.xlsx' %(expInfo['1 SubjectNumber'],expInfo['Date'],suffix)
    wb = Workbook()
    ws = wb.active
    ws.title = expName
    ws['A1'] = 'BlockType'
    ws['B1'] = 'VAS1'
    ws['C1'] = 'VAS2'
    ws['D1'] = 'VAS3'
    ws['E1'] = 'RTvas1'
    ws['F1'] = 'RTvas2'
    ws['G1'] = 'RTvas3'
    ws['H1'] = 'TimeShownVAS1'
    ws['I1'] = 'TimeShownVAS2'
    ws['J1'] = 'TimeShownVAS3'
    ws['K1'] = 'HandUsed'
    for idx,blockType in enumerate(blockTypes):
        ws['A'+str(idx+2)] = blockType
        ws['B'+str(idx+2)] = vas1[idx]
        ws['C'+str(idx+2)] = vas2[idx]
        ws['D'+str(idx+2)] = vas3[idx]
        ws['E'+str(idx+2)] = rts1[idx]
        ws['F'+str(idx+2)] = rts2[idx]
        ws['G'+str(idx+2)] = rts3[idx]
        ws['H'+str(idx+2)] = vas1t[idx]
        ws['I'+str(idx+2)] = vas2t[idx]
        ws['J'+str(idx+2)] = vas3t[idx]
        ws['K'+str(idx+2)] = handsused[idx]
    wb.save(filename = dest_filename)

def quit_task(suffix):
    saveExcel(BLOCKTYPES,VAS1,VAS2,VAS3,RTS1,RTS2,RTS3,VAS1T,VAS2T,VAS3T,HANDS,suffix)
    filename = 'DataTSSP' + os.path.sep + '%s_%s_TEMPS_%s.csv' %(expInfo['1 SubjectNumber'],expInfo['Date'],suffix)
    with open(filename, 'w') as f:
        write = csv.writer(f)
        write.writerows(all_temps)
    filename = 'DataTSSP' + os.path.sep + '%s_%s_TIMES_%s.csv' %(expInfo['1 SubjectNumber'],expInfo['Date'],suffix)
    with open(filename, 'w') as f:
        write = csv.writer(f)
        write.writerows(TOTALTIMESBLOCK)
    send_data('t')
    core.wait(.1)
    send_data(str(35))
    core.wait(.5)
    if ser!=False:
        ser.close()
    core.wait(.05)
    win.close()
    core.quit()

# initial gui
expInfo = {'1 SubjectNumber':''}

infoDlg = gui.DlgFromDict(dictionary=expInfo, title= expName)
if not infoDlg.OK:
    print('User Cancelled')

expInfo['Date'] = data.getDateStr()

# create a folder to save the data
if not os.path.isdir('DataTSSP'):
    os.makedirs('DataTSSP')


# window and text stimuli
win = visual.Window((800, 800), fullscr = FULLSCR,color = 'black')
myMouse = event.Mouse()
instru_text = u'Coloque o estimulador na sua mão '
instruct = visual.TextStim(win=win,text= instru_text,font='Arial',
                          height=0.1, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, 0.5))
instruct2 = visual.TextStim(win=win,text= u'Clique com o rato para iniciar',font='Arial',
                          height=0.06, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, -0.5))
instruct3 = visual.TextStim(win=win,text= u'descanso\n\nCarregar na tecla de espaço para continuar',font='Arial',
                          height=0.06, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, 0.5))
instruct4 = visual.TextStim(win=win,text= u'Temp ready',font='Arial',
                          height=0.06, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, 0.5))

instructVAS = visual.TextStim(win=win,text= u'Avaliar os níveis de dor',font='Arial',
                          height=0.06, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, 0.5))
ratingScale1 = visual.RatingScale(win, low=0, high=100,precision=1,showValue =False,acceptText='Aceitar',tickMarks = [0,10,20,30,40,50,60,70,80,90,100],
                                    labels=['0','10','20', '30', '40', '50', '60', '70', '80', '90', '100'], size=1.2, textSize = .8,
                                    scale=' ',pos=(0.0, -0.1),
                                    acceptPreText='Aceitar')
fixation = visual.TextStim(win=win,text= u'+',font='Arial',
                          height=0.3, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, 0.0))





# create the list of stimuli and the hand to use. Counterbalancing with 2 different orders
if int(expInfo['1 SubjectNumber'])%2==0:
    blocks =    ['CT','VI','TS']

    hands =     ['R','L','R']
else:
    blocks =    ['TS','VI','CT']
    hands =     ['L','R','L']

# Define variables to save data
BLOCKTYPES = []
VAS1 = []
VAS2 = []
VAS3 = []
RTS1 = []
RTS2 = []
RTS3 = []
VAS1T = []
VAS2T = []
VAS3T = []
TOTALTIMESBLOCK = []
HANDS = []


# send some data by serial port to test if it is alive
send_data('a')
core.wait(.6)
send_data('h')
core.wait(.6)

# define the clocks of the task
blockClock = core.Clock()
globalClock = core.Clock()
breakClock = core.Clock()

# this loop is for cleaning previous input from the serial
for i in range(100):
    serOut = read_data()

# set vibration and temperature intensity by serial port
core.wait(.1)
send_data('Z')
core.wait(.1)
send_data(str(vibration))
core.wait(.1)
send_data('t')
core.wait(.1)
send_data(str(temp))
while True:
    serOut = read_data()
    if serOut:
        instruct4.setText(serOut)
        instruct4.draw()
        win.flip()
#        if serOut[0] =='R': # if we recieve the code 'R' means that temp is ready
#            break
        try:
            r_temp = float(serOut)
            if r_temp>(temp-0.4) and r_temp<(temp+0.4):
                break
        except: core.wait(.001)
    core.wait(.3)
# show that temperature is ready
instruct4.setText('Temps ready')
instruct4.draw()
win.flip()
core.wait(2)

event.Mouse(visible = False)

# start the main loop
for idx, block in enumerate(blocks):

    # indicate hand for the heat
    if hands[idx]=='L':
        instruct.setText(instru_text + 'direita')
    else:
        instruct.setText(instru_text + 'esquerda')
    instruct.draw()
    instruct2.draw()
    win.flip()
    clear_all_cue()

    # wait until participant is ready (already saving temps)
    cont = 1
    tempcont = 1
    blockClock.reset()
    while True:
        t1 = blockClock.getTime()
        resp = get_response()
        if resp == 1:
            break
        elif resp == 99: # if 'escape' is pressed quit and save
            quit_task('TRAINING')
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()
        core.wait(.005)

    fixation.draw()
    win.flip()
    clear_all_cue()
    BLOCKTYPES.append(block)
    HANDS.append(hands[idx])
    blockStimTimes = []

    # wait 5 seconds for the participant to acommodate the hand
    t2 = blockClock.getTime()
    while t1-t2<5:
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()

    # show Heat 1
    send_data(vas_serial)
    send_trigger(11,parallel_port,parallel)
    blockStimTimes.append(globalClock.getTime())
    rtime = random.random() # wait 1.5+(0-1) seconds after the heat
    t2 = blockClock.getTime()
    blockStimTimes.append(globalClock.getTime())
    while t1-t2<rtime+1.5:
        t1 = blockClock.getTime()
        core.wait(.05)
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()

    # show VAS 1
    event.Mouse(visible = True)
    send_trigger(1,parallel_port,parallel)
    VAS1T.append(globalClock.getTime())
    while ratingScale1.noResponse:
        ratingScale1.draw()
        instructVAS.draw()
        win.flip()
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()
    RTS1.append(globalClock.getTime())
    VAS1.append(ratingScale1.getRating())
    ratingScale1.reset()
    event.Mouse(visible = False)
    fixation.draw()
    win.flip()

    # wait 5 seconds to acommodate the hand after the VAS
    t2 = blockClock.getTime()
    while t1-t2<5:
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()

    if block == 'TS': # TSSP
        send_data(ts_serial)
        send_trigger(ts_trig,parallel_port,parallel)
        trig = ts_trig
    elif block == 'VI': # TSSP + Vibration
        send_data(vi_serial)
        send_trigger(vi_trig,parallel_port,parallel)
        trig = vi_trig
    elif block == 'CT': # TSSP + CT stroking
        send_data(ct_serial)
        send_trigger(ct_trig,parallel_port,parallel)
        trig = ct_trig

    # wait 33 seconds (sending 1 trigg to the EEG 3 seconds)
    blockClock.reset()
    cont = 1
    tempcont = 1
    t1 = blockClock.getTime()
    while t1<29.998:
        t1 = blockClock.getTime()
        if t1>trial_dur*cont-0.002: # send trigg each 3 secs
            send_trigger(trig+cont,parallel_port,parallel)
            blockStimTimes.append(globalClock.getTime())
            cont+=1
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()
        core.wait(.002)

    # show Heat 12
    send_data(vas_serial)
    send_trigger(12,parallel_port,parallel)
    t1 = blockClock.getTime()
    blockStimTimes.append(globalClock.getTime())
    while t1<rtime+31.5:
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()
        core.wait(.05)

    # show VAS 2
    event.Mouse(visible = True)
    send_trigger(2,parallel_port,parallel)
    VAS2T.append(globalClock.getTime())
    while ratingScale1.noResponse:
        ratingScale1.draw()
        instructVAS.draw()
        win.flip()
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()
    RTS2.append(globalClock.getTime())
    VAS2.append(ratingScale1.getRating())
    ratingScale1.reset()
    fixation.draw()
    win.flip()
    event.Mouse(visible = False)

    t1 = blockClock.getTime()
    while t1<45:
        core.wait(.005)
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()

    # presentar VAS 3
    event.Mouse(visible = True)
    send_trigger(3,parallel_port,parallel)
    VAS3T.append(globalClock.getTime())
    while ratingScale1.noResponse:
        ratingScale1.draw()
        instructVAS.draw()
        win.flip()
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()
    RTS3.append(globalClock.getTime())
    VAS3.append(ratingScale1.getRating())
    ratingScale1.reset()
    TOTALTIMESBLOCK.append([block,blockStimTimes])
    event.Mouse(visible = False)
    fixation.draw()
    win.flip()
    t1 = blockClock.getTime()
    while t1<55:
        core.wait(.005)
        t1 = blockClock.getTime()
        if t1>temp_fs*tempcont: # read temp each 500 ms
            tempcont+=1
            read_temp()
    # BREAK
    if idx == 11:
        instruct3.draw()
        win.flip()
        cont = 1
        tempcont = 1
        blockClock.reset()
        t1 = blockClock.getTime()
        while True:
            t1 = blockClock.getTime()
            resp = get_response()
            if resp == 1:
                break
            elif resp == 99: # if scape is pressed
                quit_task ('TRAINING')
            if t1>temp_fs*tempcont: # read temp each 500 ms
                tempcont+=1
                read_temp()
            core.wait(.005)
        fixation.draw()
        win.flip()
        clear_all_cue()


instruct3.setText('OBRIGADO')
instruct3.draw()
win.flip()

core.wait(2)

quit_task('TRAINING')
