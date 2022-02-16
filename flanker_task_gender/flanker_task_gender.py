#!/usr/bin/env python
# -*- coding: utf-8 -*-


"""
EEG Flankers task to run in Psychopy v1.90.3
The task looks for a parallel port to send triggers
It also looks if a cedrus response box is avaliable, if not, responses should
be given in keyboard's space bar.
Press esc to exit from the task
Relevant data is stored in an excel file inside "data_gender" folder
"""

from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
from psychopy import visual, core, data, event, gui, parallel
import os, sys
from numpy.random import random, shuffle, choice
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import pyxid

# Turn fullscreen off for testing on monitor
FULL_SCREEN = True

# parameters that determine the behavior of the task
STIM_TIME = .75 #time of the stimulus
ISI_TIME = 1 # inter-stimulus interval
ISI_RND = .5
CF_TRIALS = 60 # number congruent female trials
IF_TRIALS = 60 # number incongruent female trials
SF_TRIALS = 30 # number of scrambled female trials
CM_TRIALS = 60 # number congruent male trials
IM_TRIALS = 60 # number incongruent male trials
SM_TRIALS = 30 # number of scrambled male trials

task_instructions = u'Responde ao xénero da cara central\n"z" para home; "m" para muller.\n\nPreme a barra espaciadora para comezar'
rest_instructions = u'Descanso.\n\nPreme a barra espaciadora para continuar'
thanks_text = u'Obrigado'

# create trial type list
TrialType =[1] * CF_TRIALS
TrialType.extend([2] * CM_TRIALS)
TrialType.extend([3] * IF_TRIALS)
TrialType.extend([4] * IM_TRIALS)
TrialType.extend([5] * SF_TRIALS)
TrialType.extend([6] * SM_TRIALS)
shuffle(TrialType)


path = os.getcwd() + os.path.sep + 'gender_faces'
stims= [f for f in os.listdir(path) if f.endswith('.jpg')]

# create_trials
path = os.getcwd() + os.path.sep + 'gender_faces' + os.path.sep + 'Woman_central'
FE_C= [f for f in os.listdir(path) if f.endswith('.jpg')]
path = os.getcwd() + os.path.sep + 'gender_faces' + os.path.sep + 'Man_central'
MA_C= [f for f in os.listdir(path) if f.endswith('.jpg')]
path = os.getcwd() + os.path.sep + 'gender_faces' + os.path.sep + 'Woman_flankers'
FE_F= [f for f in os.listdir(path) if f.endswith('.jpg')]
path = os.getcwd() + os.path.sep + 'gender_faces' + os.path.sep + 'Man_flankers'
MA_F= [f for f in os.listdir(path) if f.endswith('.jpg')]
path = os.getcwd() + os.path.sep + 'gender_faces' + os.path.sep + 'Woman_scrambled'
FE_S= [f for f in os.listdir(path) if f.endswith('.jpg')]
path = os.getcwd() + os.path.sep + 'gender_faces' + os.path.sep + 'Man_scrambled'
MA_S= [f for f in os.listdir(path) if f.endswith('.jpg')]

# shuffle stims
FE_C2 = FE_C[:]
FE_F2 = FE_F[:]
MA_C2 = MA_C[:]
MA_F2 = MA_F[:]
FE_C3 = FE_C[:]
FE_F3 = FE_F[:]
MA_C3 = MA_C[:]
MA_F3 = MA_F[:]
FE_C4 = FE_C[:]
MA_C4 = MA_C[:]
FE_S2 = FE_S[:]
MA_S2 = MA_S[:]
shuffle(FE_C2)
shuffle(FE_F2)
shuffle(MA_C2)
shuffle(MA_F2)
shuffle(FE_C3)
shuffle(FE_F3)
shuffle(MA_C3)
shuffle(MA_F3)
shuffle(FE_C4)
shuffle(MA_C4)
shuffle(FE_S2)
shuffle(MA_S2)

# append stims
trials = []
for trial in TrialType:
    if trial == 1: # congruent female
        trials.append([FE_C2.pop(), FE_F2.pop()])
    elif trial == 2: # congruent male
        trials.append([MA_C2.pop(), MA_F2.pop()])
    elif trial == 3: # incongruent female
        trials.append([FE_C3.pop(), MA_F3.pop()])
    elif trial == 4: # incongruent male
        trials.append([MA_C3.pop(), FE_F3.pop()])
    elif trial == 5: # scrambled female
        trials.append([FE_C4.pop(), FE_S2.pop()])
    elif trial == 6: # scrambled male
        trials.append([MA_C4.pop(), MA_S2.pop()])

def send_trigger(code,parallel_port,parallel):
    if parallel_port == 1:
        parallel.setData(code)
        core.wait(.005)
        parallel.setData(0)
    else:
        pass


# clear response queue
def clear_all_queue(dev,cedrus):
    if cedrus:
        while True:
            dev.poll_for_response()
            if dev.response_queue_size():
                dev.clear_response_queue()
            else: break
    else:
        while True:
            if event.getKeys(): event.clearEvents('keyboard')
            else: break

# get response from cedrus, or the keyboard in its absence
def get_response(cedrus,dev):
    key_out=[]
    if cedrus:
        dev.poll_for_response()
        if dev.response_queue_size()>0:
            resp = dev.get_next_response()
            if resp['pressed']==True:
                #key = resp['key']
                key = 1
                return key
        else:
            resp=event.getKeys()
            if resp:
                if resp[0]=='escape':key_out=99
                return key_out
    else:
        resp=event.getKeys()
        if resp:
            key=resp[0]
            if key=='space':key_out=9#code that sends for spacebar
            elif key=='z':key_out=11#code that sends for z
            elif key=='m':key_out=12#code that sends for m
            elif key=="escape":key_out=99
            return key_out

#parallel port setup
try:
    parallel.setPortAddress(0x378)#direccion do porto paralelo
    parallel.setData(0)
    parallel_port=1
except:
    parallel_port=0
    print "Oops!  No parallel port found"


# Cedrus response box setup
try:
    devices = pyxid.get_xid_devices()
    dev = devices[0] # response box will be "dev"
    if dev.is_response_device():
        dev.reset_base_timer()
        dev.reset_rt_timer()
    dev.clear_response_queue()
    cedrus=1
    print "Cedrus response box found! Maybe you need to change the response keys (see line 121)"

except:
    print "No cedrus response box found, we will use the keyboard"
    cedrus=0
    dev=None


# Store info about the experiment session
expName = 'FlankerGender'
# gui dialogue to get participant id
expInfo = {'ID Participant':''}
dlg = gui.DlgFromDict(dictionary=expInfo, title=expName, fixed=[])
# if user pressed cancel, quit
if dlg.OK == False:
    core.quit()
# more configuration parameters
expInfo['date'] = data.getDateStr()  # timestamp
expInfo['expName'] = expName

# Setup files for saving
if not os.path.isdir('data_gender'):
    os.makedirs('data_gender')
filename = 'data_gender' + os.path.sep + '%s_%s' %(expInfo['ID Participant'], expInfo['date'])

# EXCEL
xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]
xlSheet.title = 'Flanker' + str(expInfo['ID Participant']) + str(expInfo['date'])

xlSheet.cell('A1').value = 'stimType 1:con Fem 2:con Male 3:inc Fem 4:inc Male 5:scr Fem 6:scr Male'
xlSheet.cell('B1').value = 'CentralStim'
xlSheet.cell('C1').value = 'FlankerStim'
xlSheet.cell('D1').value = 'StimDuration'
xlSheet.cell('E1').value = 'RT'
xlSheet.cell('F1').value = 'Response'
xlSheet.cell('G1').value = 'TotalTime'

# Setup the Window
win = visual.Window(size=(800, 768), fullscr=FULL_SCREEN, units='norm')

# display components
instruct_text = visual.TextStim(win=win, ori=0, name='instruct_text', text=task_instructions,
            font='Arial',alignHoriz='center', alignVert='center', height=0.1, wrapWidth=1.5, color='white')
fixation = visual.ShapeStim(win,units='cm', lineColor='white', lineWidth=1.2, vertices=((-0.9, 0), (0.9, 0), (0,0), (0,0.9), (0,-0.9)),
            interpolate=True, closeShape=False, pos=(0,0))
pic1 = visual.ImageStim(win, image='gender_faces' + os.path.sep + 'Woman_central'+ os.path.sep + '022_y_f_n_a_RS.jpg', pos=(-0.8, 0.0))
pic2 = visual.ImageStim(win, image='gender_faces' + os.path.sep + 'Woman_central'+ os.path.sep + '022_y_f_n_a_RS.jpg', pos=(-0.4, 0.0))
pic3 = visual.ImageStim(win, image='gender_faces' + os.path.sep + 'Woman_central'+ os.path.sep + '022_y_f_n_a_RS.jpg', pos=(0.4, 0.0))
pic4 = visual.ImageStim(win, image='gender_faces' + os.path.sep + 'Woman_central'+ os.path.sep + '022_y_f_n_a_RS.jpg', pos=(0.8, 0.0))
pic5 = visual.ImageStim(win, image='gender_faces' + os.path.sep + 'Woman_central'+ os.path.sep + '022_y_f_n_a_RS.jpg', pos=(0.0, 0.0))
event.Mouse(visible=False)
thanks = visual.TextStim(win=win, ori=0, name='thanks', text=thanks_text, font='Arial', height=0.1, color='white')


# Initialize timers
globalClock = core.Clock()  # to track the time since experiment started
trialClock = core.Clock()

# task starts
instruct_text.draw()
win.flip()
core.wait(1)
clear_all_queue(dev, cedrus)
while True:
    resp = get_response(cedrus,dev)
    if resp==9:
        break
win.flip()
core.wait(1)
clear_all_queue(dev, cedrus)
nEnsaio = 1
globalClock.reset()
blockTrial = 1
for idx, trial in enumerate(trials):

    if TrialType[idx] == 1 or TrialType[idx] == 4:
        pic1.setImage('gender_faces' + os.path.sep + 'Woman_flankers'+ os.path.sep + trial[1])
        pic2.setImage('gender_faces' + os.path.sep + 'Woman_flankers'+ os.path.sep + trial[1])
        pic3.setImage('gender_faces' + os.path.sep + 'Woman_flankers'+ os.path.sep + trial[1])
        pic4.setImage('gender_faces' + os.path.sep + 'Woman_flankers'+ os.path.sep + trial[1])
    elif TrialType[idx] == 2 or TrialType[idx] == 3:
        pic1.setImage('gender_faces' + os.path.sep + 'Man_flankers'+ os.path.sep + trial[1])
        pic2.setImage('gender_faces' + os.path.sep + 'Man_flankers'+ os.path.sep + trial[1])
        pic3.setImage('gender_faces' + os.path.sep + 'Man_flankers'+ os.path.sep + trial[1])
        pic4.setImage('gender_faces' + os.path.sep + 'Man_flankers'+ os.path.sep + trial[1])
    elif TrialType[idx] == 5:
        pic1.setImage('gender_faces' + os.path.sep + 'Woman_scrambled'+ os.path.sep + trial[1])
        pic2.setImage('gender_faces' + os.path.sep + 'Woman_scrambled'+ os.path.sep + trial[1])
        pic3.setImage('gender_faces' + os.path.sep + 'Woman_scrambled'+ os.path.sep + trial[1])
        pic4.setImage('gender_faces' + os.path.sep + 'Woman_scrambled'+ os.path.sep + trial[1])
    elif TrialType[idx] == 6:
        pic1.setImage('gender_faces' + os.path.sep + 'Man_scrambled'+ os.path.sep + trial[1])
        pic2.setImage('gender_faces' + os.path.sep + 'Man_scrambled'+ os.path.sep + trial[1])
        pic3.setImage('gender_faces' + os.path.sep + 'Man_scrambled'+ os.path.sep + trial[1])
        pic4.setImage('gender_faces' + os.path.sep + 'Man_scrambled'+ os.path.sep + trial[1])

    if TrialType[idx] == 1 or TrialType[idx] == 3 or TrialType[idx] == 5:
        pic5.setImage('gender_faces' + os.path.sep + 'Woman_central'+ os.path.sep + trial[0])
    else:
        pic5.setImage('gender_faces' + os.path.sep + 'Man_central'+ os.path.sep + trial[0])
    pic1.draw()
    pic2.draw()
    pic3.draw()
    pic4.draw()
    pic5.draw()
    win.flip()
    trialClock.reset()
    send_trigger(TrialType[idx],parallel_port,parallel)
    clear_all_queue(dev,cedrus)
    respondeu = 0
    while trialClock.getTime()<STIM_TIME-.01:
        if not respondeu:
            resp = get_response(cedrus,dev) #if responded
            if resp:
                if resp==11:
                    rt = trialClock.getTime()
                    send_trigger(resp,parallel_port,parallel)
                    respondeu = 1
                elif resp==12:
                    rt = trialClock.getTime()
                    send_trigger(resp,parallel_port,parallel)
                    respondeu = 1
                elif resp==99:
                    xlBook.save(filename + '_inc.xlsx')
                    win.close()
                    core.quit()
    else:
        core.wait(.002)
    win.flip()
    duration = trialClock.getTime()
    rnd_isi = random()*ISI_RND
    while trialClock.getTime()<ISI_TIME+STIM_TIME+rnd_isi:
        if not respondeu:
            resp = get_response(cedrus,dev) #if responded
            if resp:
                if resp==11:
                    rt = trialClock.getTime()
                    send_trigger(resp,parallel_port,parallel)
                    respondeu = 1
                elif resp==12:
                    rt = trialClock.getTime()
                    send_trigger(resp,parallel_port,parallel)
                    respondeu = 1
                elif resp==99:
                    xlBook.save(filename + '_inc.xlsx')
                    win.close()
                    core.quit()
    else:
        core.wait(.002)

    # save trial data to excel
    xlSheet.cell('A'+str(idx+2)).value = TrialType[idx]
    xlSheet.cell('B'+str(idx+2)).value = trial[0]
    xlSheet.cell('C'+str(idx+2)).value = trial[1]
    xlSheet.cell('D'+str(idx+2)).value = duration
    if respondeu:
        xlSheet.cell('E'+str(idx+2)).value = rt
        xlSheet.cell('F'+str(idx+2)).value = resp
    xlSheet.cell('G'+str(idx+2)).value = globalClock.getTime()

    # rest period
    if idx ==149:
        instruct_text.setText(rest_instructions)
        instruct_text.draw()
        win.flip()
        core.wait(1)
        clear_all_queue(dev, cedrus)
        while True:
            resp = get_response(cedrus,dev)
            if resp == 9:
                break
        win.flip()


thanks.draw()
win.flip()
core.wait(3)
xlBook.save(filename + '.xlsx')
win.close()
core.quit()
