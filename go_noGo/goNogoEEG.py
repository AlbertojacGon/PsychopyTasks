#!/usr/bin/env python
# -*- coding: utf-8 -*-


"""
EEG Go-noGo task to run in Psychopy 1.83
The task looks for a parallel port to send triggers
It also looks if a cedrus response box is avaliable, if not, responses should
be given in keyboard's space bar.
Press esc to exit from the task
Relevant data is stored in an excel file inside "data" folder
"""

from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
from psychopy import visual, core, data, event, gui, parallel
import os, sys
from numpy.random import  random, shuffle
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import pyxid


# Turn fullscreen off for testing on monitor
FULL_SCREEN = True

# parameters that determine the behavior of the task
STIM_TIME = 1.5 #time of the stimulus
STIM_ISI_SECONDS = 1.5 # inter-stimulus interval
TRIALS = 300 # number of trials
PROP_GO = .8 # proportion of go trials
N_BLOCKS = 10 # number of blocks
REST_TIME = 60 # resting seconds between blocks

N_GO = int(TRIALS*PROP_GO)
N_NOGO = int(round(TRIALS*(1-PROP_GO)))
trialsXblock = TRIALS/N_BLOCKS

def send_trigger(code,parallel_port,parallel):
    if parallel_port == 1:
        parallel.setData(code)
    else:
        pass


# funcion para borrar a cola de respostas
def clear_all_queue(dev,cedrus):
    if cedrus:
        while True:
            dev.poll_for_response()
            if dev.response_queue_size():
                dev.clear_response_queue()
            else: break
    else:
        while True:
            if event.getKeys(): event.clearEvents('keyboard')
            else: break

def get_response(cedrus,dev):
    key_out=[]
    if cedrus:
        dev.poll_for_response()
        if dev.response_queue_size()>0:
            resp = dev.get_next_response()
            if resp['pressed']==True:
                #key = resp['key']
                key = 1
                return key
        else:
            resp=event.getKeys()
            if resp:
                if resp[0]=='escape':key_out=99
                return key_out
    else:
        resp=event.getKeys()
        if resp:
            key=resp[0]
            if key=='space':key_out=1#codigo que deberia mandar o cedrus lumina400pair ao pulsar o segundo botón da man esquerda
            elif key=="escape":key_out=99
            return key_out

#Direccion do porto paralelo
try:
    parallel.setPortAddress(0x378)#direccion do porto paralelo
    parallel.setData(0)
    parallel_port=1
except:
    parallel_port=0
    print "Oops!  No parallel port found"


# Cedrus response box
try:
    devices = pyxid.get_xid_devices()
    dev = devices[0] # response box will be "dev"
    if dev.is_response_device():
        dev.reset_base_timer()
        dev.reset_rt_timer()
    dev.clear_response_queue()
    cedrus=1
    print "Cedrus response box found! Maybe you need to change the response keys (see line 121)"

except:
    print "No cedrus response box found, we will use the keyboard"
    cedrus=0
    dev=None




task_instructions = u'press the space bar as fast as you can when O appears \nPress the space bar to begin'


stimsGo = ['o' for i in range(N_GO)]
stimsNogo = ['x' for i in range(N_NOGO)]
stims = stimsGo + stimsNogo
shuffle(stims)

# Store info about the experiment session
expName = 'GoNogo'

# gui dialogue to get participant id
expInfo = {'ID Participant':''}

dlg = gui.DlgFromDict(dictionary=expInfo, title=expName, fixed=[])

# if user pressed cancel, quit
if dlg.OK == False:
    core.quit()

# set a few more configuration parameters
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName


# Setup files for saving
if not os.path.isdir('data'):
    # if this fails (e.g. permissions) we will get error
    os.makedirs('data')

filename = 'data' + os.path.sep + '%s_%s' %(expInfo['ID Participant'],\
            expInfo['date'])



# EXCEL

xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]
xlSheet.title = 'MSIT' + str(expInfo['ID Participant']) + str(expInfo['date'])

xlSheet.cell('A1').value = 'stim'
xlSheet.cell('B1').value = '1 = Go; 0 = NoGo'
xlSheet.cell('C1').value = 'Stim Duration'
xlSheet.cell('D1').value = 'RT'
xlSheet.cell('E1').value = '0=Incorrect; 1=Correct'
xlSheet.cell('F1').value = 'TotalTime'


# Setup the Window
win = visual.Window(size=(800, 768),
                    fullscr=FULL_SCREEN,
                    screen=0,
                    allowGUI=False,
                    allowStencil=False,
                    monitor='testMonitor',
                    color='black',
                    colorSpace='rgb')




# display components
instruct_text = visual.TextStim(win=win, ori=0, name='instruct_text',
    text=task_instructions,
    font='Arial',alignHoriz='center', alignVert='center',
    pos=[0, 0], height=0.1, wrapWidth=1.5,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)

fixation = visual.ShapeStim(win,units='cm', lineColor='white',
    lineWidth=1.2, vertices=((-0.9, 0), (0.9, 0), (0,0), (0,0.9), (0,-0.9)),
    interpolate=True, closeShape=False, pos=(0,0))

stimuli = visual.TextStim(win=win, text='', height=0.3)

thanks = visual.TextStim(win=win, ori=0, name='thanks',
    text='Obrigado',    font='Arial',
    pos=[0, 0], height=0.1, wrapWidth=1,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)


# Initialize timers
globalClock = core.Clock()  # to track the time since experiment started
trialClock = core.Clock()


# task starts

instruct_text .draw()
win.flip()
core.wait(2)
clear_all_queue(dev, cedrus)
while True:
    resp = get_response(cedrus,dev)
    if resp:
        core.wait(1)
        break

clear_all_queue(dev, cedrus)
nEnsaio = 1
globalClock.reset()
blockTrial = 1
for stim in stims:
    if blockTrial   == trialsXblock:
        blockTrial = 1
        instruct_text.setText('Break')
        instruct_text.draw()
        win.flip()
        core.wait(REST_TIME)
        clear_all_queue(dev, cedrus)
        instruct_text.setText('Press the space bar to continue')
        instruct_text.draw()
        win.flip()
        while True:
            resp = get_response(cedrus,dev)
            if resp:
                core.wait(1)
                break

    blockTrial += 1
    clear_all_queue(dev, cedrus)
    if stim == 'o':
        stimuli.setText(stim)
        code = 1
    else:
        stimuli.setText(stim)
        code = 2
    stimuli.draw()
    rt = None
    respondeu = 0
    win.flip()
    trialClock.reset()
    send_trigger(code,parallel_port,parallel)
    clear_all_queue(dev,cedrus)
    core.wait(.01)
    send_trigger(0,parallel_port,parallel)
    while trialClock.getTime()<STIM_TIME-(.005):
        if not respondeu:
            resp = get_response(cedrus,dev) #if responded
            if resp:
                if resp==1:
                    rt = trialClock.getTime()
                    send_trigger(9,parallel_port,parallel)
                    respondeu = 1
                if resp==99:
                    xlBook.save(filename + '_inc.xlsx')
                    win.close()
                    core.quit()
        else:
            core.wait(.01)
    fixation.draw()
    win.flip()
    t_stim = trialClock.getTime()
    while trialClock.getTime()<(STIM_TIME +STIM_ISI_SECONDS) -(.005):
        if not respondeu:
            resp = get_response(cedrus, dev)
            if resp:
                if resp==1:
                    rt = trialClock.getTime()
                    send_trigger(9,parallel_port,parallel)
                    respondeu = 1
                if resp==99:
                    xlBook.save(filename + '_inc.xlsx')
                    win.close()
                    core.quit()
        else:
            core.wait(.01)
    clear_all_queue(dev, cedrus)
    send_trigger(0,parallel_port,parallel)
    trialClock.reset()
    if stim =='x':
        typeStim = 0
    else:
        typeStim = 1
    if stim == 'o' and rt:
        correction = 1
    elif stim == 'x' and not rt:
        correction = 1
    else: correction = 0

    xlSheet.cell('A'+str(nEnsaio+1)).value = stim
    xlSheet.cell('B'+str(nEnsaio+1)).value = typeStim
    xlSheet.cell('C'+str(nEnsaio+1)).value = t_stim
    xlSheet.cell('D'+str(nEnsaio+1)).value = rt
    xlSheet.cell('E'+str(nEnsaio+1)).value = correction
    xlSheet.cell('F'+str(nEnsaio+1)).value = globalClock.getTime()
    nEnsaio+=1



thanks.draw()
win.flip()
core.wait(3)
xlBook.save(filename + '.xlsx')
win.close()
core.quit()
