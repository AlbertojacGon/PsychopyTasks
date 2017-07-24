#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
fMRI Go-noGo task to run in Psychopy 1.83
The task waits for mouse left click to start (scan trigger in the cedrus lumina).
Presentation times are extracted from "times.par" file
Relevant data is stored in an excel file
"""
from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
from psychopy import visual, core, data, event, gui
import os, sys, csv
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook



# Turn fullscreen off for testing on monitor
FULL_SCREEN= True


task_instructions1 = u'press a button as fast as you can only when "O" appears'

begin_text = u'The task is going to start'
thanks_text = u'Thankyou'

# read file with onset times

with open('times.par', 'rb') as csvfile:
    listStim = []
    listTime = []
    spamreader = csv.reader(csvfile, delimiter= '\t')
    for row in spamreader:
        listStim.append(row[0][12:13])
        listTime.append(row[0][17:20])

listTime = [float(i) for i in listTime]
listStim = [int(i) for i in listStim]

# clear all responses
def clear_all_queue():
    while True:
        if event.getKeys(): event.clearEvents('keyboard')
        else: break

# get responses from the cedrus lumina fMRI controller (in keyboard mode)
def get_response():
    key_out=[]
    resp=event.getKeys()
    if resp:
        key=resp[0]
        if key=='lshift':key_out=98 #scanner trigger will be 98
        elif key=='a':key_out=1 # al responses will be 1
        elif key=='b':key_out=1
        elif key=='c':key_out=1
        elif key=='d':key_out=1
        elif key=="escape":key_out=99
        return key_out




# Store info about the experiment session
expName = 'GoNogo_fMRI'

# gui dialogue to get participant id and group
expInfo = {'ID_Participant':'', 'Group':['CG','Exp']}


dlg = gui.DlgFromDict(dictionary=expInfo, title=expName, fixed=[])

# if user pressed cancel, quit
if dlg.OK == False:
    core.quit()

# set a few more configuration parameters
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName




# Set up output files

# Setup files for saving
if not os.path.isdir('data'+ os.path.sep):
    # if this fails (e.g. permissions) we will get error
    os.makedirs('data'+ os.path.sep )

filename = 'data' + os.path.sep + '%s%s_%s' %(expInfo['ID_Participant'],\
                                    expInfo['Group'], expInfo['date'])



# Excel file for saving

xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]
xlSheet.title = 'GoNogo'

xlSheet.cell('A1').value = 'stim'
xlSheet.cell('B1').value = 'RT'
xlSheet.cell('C1').value = 'Correction'
xlSheet.cell('D1').value = 'stimOnset'
xlSheet.cell('E1').value = 'nullOnset'
xlSheet.cell('F1').value = 'nullDuration'


# Setup the Window
win = visual.Window(size=(800, 768),
                    fullscr=FULL_SCREEN,
                    screen=0,
                    allowGUI=False,
                    allowStencil=False,
                    monitor='testMonitor',
                    color='black',
                    colorSpace='rgb')



# Initialize stimulus
instruct_text = visual.TextStim(win=win, ori=0, name='instruct_text',
    text=task_instructions1,
    font='Arial',alignHoriz='center', alignVert='center',
    pos=[0, 0], height=0.1, wrapWidth=1.5,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)

beginText = visual.TextStim(win=win, ori=0, name='begin_text',
    text=begin_text,
    font='Arial',alignHoriz='center', alignVert='center',
    pos=[0, 0], height=0.1, wrapWidth=1.5,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)

fixation = visual.ShapeStim(win,units='cm', lineColor='white',
    lineWidth=1.2, vertices=((-0.5, 0), (0.5, 0), (0,0), (0,0.5), (0,-0.5)),
    interpolate=True, closeShape=False, pos=(0,0))
goStim = visual.TextStim(win=win, text='o', height=0.3)
stStim = visual.TextStim(win=win, text='x', height=0.3)

thanks = visual.TextStim(win=win, ori=0, name='thanks',
    text=thanks_text,    font='Arial',
    pos=[0, 0], height=0.1, wrapWidth=1,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)


# Initialize timers
globalClock = core.Clock()  # to track the time since experiment started
trialClock = core.Clock()

trial = 0
totalTime = 0
instruct_text .draw()
win.flip()
clear_all_queue()
while True:
    resp = get_response()
    if resp:
        if resp == 98:
            globalClock.reset()
            trialClock.reset()
            break

beginText.draw()
win.flip()
while globalClock.getTime()< 6:
    resp = get_response()
    if resp:
        if resp == 99:
            win.close()
            core.quit()
            break

win.flip()
while globalClock.getTime()< 7.99:
    resp = get_response()
    if resp:
        if resp == 99:
            win.close()
            core.quit()
            break
totalTime += 8
clear_all_queue()

for idx,stim in enumerate(listStim):
    null = 0
    if stim == 0:
        fixation.draw()
        win.flip()
        nullTime = globalClock.getTime()

    elif stim == 1:
        goPic.draw()
        win.flip()

    elif stim == 2:
        stStim.draw()
        win.flip()
    if stim:
        trialClock.reset()
        showTime = globalClock.getTime()
        clear_all_queue()
        trial += 1
        rt = None
        responded = 0

    trialTime = listTime[idx]
    totalTime += trialTime

    while globalClock.getTime() < totalTime -.01:
        if not responded:
            resp = get_response()
            if resp:
                if resp==1:
                    rt = trialClock.getTime()
                    responded = 1
                if resp==99:
                    # if scape save excel with "_inc" suffix
                    xlBook.save(filename + '_inc.xlsx')
                    win.close()
                    core.quit()
        else:
            core.wait(.003)


    if stim:
        xlSheet.cell('A'+str(trial+1)).value = stim
        xlSheet.cell('D'+str(trial+1)).value = showTime
    else:
        if listStim[idx-1] == 1 and rt:
            correction = 1
        elif listStim[idx-1] == 2 and not rt:
            correction = 1
        else: correction = 0
        xlSheet.cell('B'+str(trial+1)).value = rt
        xlSheet.cell('C'+str(trial+1)).value = correction
        xlSheet.cell('E'+str(trial+1)).value = nullTime
        xlSheet.cell('F'+str(trial+1)).value = listTime[idx]



thanks.draw()
win.flip()
core.wait(3)
xlBook.save(filename + '.xlsx')
