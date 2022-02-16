# -*- coding: utf-8 -*-

from __future__ import division
from psychopy import visual, core, data, event, gui
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import os



fullScreen = True

# times in seconds
restTime = 12.5
stimTime = 2
interstimTime = 1.5


# volumes to skip during T1 stabilization
volumes2skip = 3
TR = 2

# Are we going to use lumina response box? 0 = No; 1 = Yes
LUMINA = 1

# Trigger send by the lumina
LUMINA_TRIGGER = 'lshift'


###################
# Setup da tarefa #
###################


# function to clear responses
def clear_all_cue():
    while True:
        if event.getKeys():
            event.clearEvents('keyboard')
        else: break


# funcion para recoller as respostas do teclado ou da cedrus
def get_response(luminaTrigg):
    key2=[]
    resp=event.getKeys()
    if resp:
        key=resp[0]
        if key==luminaTrigg: key2 = 1
        elif key=="escape": key2 = 9
        return key2




# make folder to save files
if not os.path.isdir('data'):
    os.makedirs('data')
carpeta = os.getcwd()+ os.path.sep


# Task info
expName=u'mother_attach_session3'
expInfo = { 'MotherID':['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16']}
dlg = gui.DlgFromDict(dictionary=expInfo, title=expName)
if dlg.OK == False: core.quit()  #clickou Cancelar

expInfo['date'] = data.getDateStr()
expInfo['expName'] = expName
filename = carpeta+ os.path.sep +'data' + os.path.sep + 'Sub%s_%sSession3' %(expInfo['MotherID'], expInfo['date'])


if int(expInfo['MotherID'])%2 == 0:
    sessions = [u'EMPATIZAR']
else:
    sessions = [u'IMITAR']


# load stims
wb = load_workbook(filename = carpeta + 'Stim_order_MotherInfantAttach.xlsx')
ws = wb.active
stims = []
for col in ws.columns:
    if col[0].value == 'Subject' + expInfo['MotherID'] + 'C':
        for cell in col:
            stims.append(cell.value)
stims = stims[37:56]



# #preparar o archivo para gardar os datos en excel
xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]
xlSheet.title = 'Subject' + expInfo['MotherID']
xlSheet.cell('A1').value = 'Stim'
xlSheet.cell('B1').value = 'onsetTime'
xlSheet.cell('C1').value = 'offsetTime'
xlSheet.cell('D1').value = 'empathy_imitation'

#ventana
win = visual.Window(size=(880, 768), fullscr=fullScreen, screen=True, allowGUI=False, allowStencil=False,#ventana
    monitor='testMonitor', color='black', colorSpace='rgb')

# punto de fixacion
fixation = visual.ShapeStim(win,units='cm', lineColor='white', lineWidth=1.2, #punto de fixacion
    vertices=((-0.5, 0), (0.5, 0), (0,0), (0,0.5), (0,-0.5)), interpolate=False,
    closeShape=False, pos=(0,0))

# relojes para comprobar que todo vaia ao seu tempo
globalClock = core.Clock()

session = sessions.pop(0)
#aqui creamos os diferentes estimulos
instruccions = visual.TextStim(win, text=session, font=u'Arial', height=.3, color=u'white', colorSpace=u'rgb')


instruccions2 = visual.TextStim(win=win, ori=0, name='a experiência vai começar', #instruccions
    text=u'a experiência vai começar', font=u'Arial',
    pos=[0, 0], height=0.1, wrapWidth=None, alignHoriz='center', alignVert='center',
    color=u'white', colorSpace=u'rgb', opacity=1,
    depth=0.0)

propS = win.size[0]/win.size[1]
picture = visual.ImageStim(win, image=None, units='norm', pos=(0.0, 0.0), size=(.85, .85*propS))

##########
# TAREFA #
##########

instruccions.draw()
win.flip()

# wait for 1st trigger
clear_all_cue()
while True:
    resp = get_response(LUMINA_TRIGGER)
    if resp == 1:
        globalClock.reset()
        win.flip()
        break
    elif resp == 9:
        win.close()
        core.quit()

# some volumes to skip for T1 stabilization
timeTotal=volumes2skip*TR
skipped_volumes = 0
while globalClock.getTime()+.05 <= volumes2skip*TR:
    resp = get_response(LUMINA_TRIGGER)
    if resp == 9:
        win.close()
        core.quit()
win.flip()

blocks = 18
trial = 1

for block in range(blocks):
    #rest
    while globalClock.getTime()+.1 <= timeTotal + restTime:
        resp = get_response(LUMINA_TRIGGER)
        if resp == 9:
            win.close()
            xlsxWriter.save(filename = ( filename+ 'inc.xlsx'))
            core.quit()
            core.wait(.01)

    timeTotal = timeTotal + restTime

    blockType = stims.pop(0)
    for imIdx in range(1,5,1):
       if blockType[-1] =='2': #in blocks 2 we show pics 5,6,7,8
           stimName = blockType[0:-1] + str(imIdx+4) + '.jpg'
           picture.setImage('Subject' + expInfo['MotherID'] + os.path.sep + stimName)
           picture.draw()
           while True:
               if globalClock.getTime()+.015 >= timeTotal:
                  win.flip()
                  onTime = globalClock.getTime()
                  break

       else: #in blocks 1 or 3 we show pics 1,2,3,4
           stimName = blockType[0:-1] + str(imIdx) + '.jpg'
           picture.setImage('Subject' + expInfo['MotherID'] + os.path.sep + stimName)
           picture.draw()
           while True:
               if globalClock.getTime()+.015 >= timeTotal:
                  win.flip()
                  onTime = globalClock.getTime()
                  break

       while globalClock.getTime()+.03 <= timeTotal + stimTime:
           resp = get_response(LUMINA_TRIGGER)
           if resp == 9:
               win.close()
               xlsxWriter.save(filename = ( filename+ 'inc.xlsx'))
               core.quit()
           core.wait(.01)
       xlSheet.cell('A'+str(trial+1)).value = stimName
       xlSheet.cell('B'+str(trial+1)).value = onTime
       xlSheet.cell('D'+str(trial+1)).value = session
       timeTotal = timeTotal + stimTime
        # interstims black screen
       while True:
           if globalClock.getTime()+.015 >= timeTotal:
               win.flip()
               offTime = globalClock.getTime()
               break
       xlSheet.cell('C'+str(trial+1)).value = offTime

       while globalClock.getTime()+.1 <= timeTotal + interstimTime:
           resp = get_response(LUMINA_TRIGGER)
           if resp == 9:
               win.close()
               xlsxWriter.save(filename = ( filename+ 'inc.xlsx'))
               core.quit()
           core.wait(.01)
       trial+=1
       timeTotal = timeTotal + interstimTime


while globalClock.getTime()+.1 <= timeTotal + restTime:
        resp = get_response(LUMINA_TRIGGER)
        if resp == 9:
            win.close()
            xlsxWriter.save(filename = ( filename+ 'inc.xlsx'))
            core.quit()
            core.wait(.01)

instruccions.setText(u'Obrigado')
instruccions.draw()
win.flip()
xlsxWriter.save(filename = ( filename+ '.xlsx'))
core.wait(1)
win.close()
core.quit()
