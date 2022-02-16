#!/usr/bin/env python
# -*- coding: utf-8 -*-


"""
EEG picture naming task to run in Psychopy 1.85.4
Press esc to exit from the task
Relevant data is stored in an excel file inside "data" folder
"""

from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
from psychopy import visual, core, data, event, gui, parallel
import os, sys
from numpy.random import  random, shuffle
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook




# Turn fullscreen off for testing on monitor that == not 1024x768
FULL_SCREEN = True

# CC here are some parameters that determine the behavior of the task
STIM_TIME = 3.0
ISI_TIME = 1.7
RND_TIME = 0.6 #random

#Direccion do porto paralelo
try:
    parallel.setPortAddress(0x378)#direccion do porto paralelo
    parallel.setData(0)
    parallel_port=1
except:
    parallel_port=0
    print "Oops!  No parallel port found"



def send_trigger(code,parallel_port,parallel):
    if parallel_port == 1:
        parallel.setData(code)
        core.wait(0.005)
        parallel.setData(0)
    else:
        pass


# clear all responses
def clear_all_cue(dev,cedrus):
    if cedrus:
        while True:
            dev.poll_for_response()
            if dev.response_queue_size():
                dev.clear_response_queue()
            else: break
    else:
        while True:
            if event.getKeys(): event.clearEvents('keyboard')
            else: break

# get responses
def get_response():
    key_out=[]
    resp=event.getKeys()
    if resp:
        key=resp[0]
        if key=="escape":key_out=99
        else:
            key_out=1
        return key_out

# clear all responses
def clear_all_cue():
    while True:
        if event.getKeys(): event.clearEvents('keyboard')
        else: break

task_instructions = u'Ready for a short practice? We will see 10 pictures for getting ready before the main test. Please look at each picture and try to say the name of each picture silently and without moving your jaw and mouth.'

expName = 'PicNaming'  # from the Builder filename that created this script

# gui dialogue to get participant id, session number, and type of first
# block (useful for counterbalancing)
expInfo = {'1. ID Participant':'', '2. Group':''}
dlg = gui.DlgFromDict(dictionary=expInfo, title=expName, fixed=[])

# if user pressed cancel, quit
if dlg.OK == False:
    core.quit()

# set a few more configuration parameters
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName


# Setup files for saving
if not os.path.isdir('data'+ os.path.sep + '%s' %(expInfo['expName']) ):
    # if this fails (e.g. permissions) we will get error
    os.makedirs('data'+ os.path.sep + '%s' %(expInfo['expName']) )

filename = 'data' + os.path.sep + '%s' %(expInfo['expName']) \
    + os.path.sep + '%s_%s_%s' %(expInfo['1. ID Participant'],expInfo['2. Group'],\
                                    expInfo['date'])



#########
# EXCEL #
#########

xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]
xlSheet.title = 'picNaming' + str(expInfo['1. ID Participant'])

xlSheet.cell('A1').value = 'Trial'
xlSheet.cell('B1').value = 'StimName'
xlSheet.cell('C1').value = 'StimOnsetTime'
xlSheet.cell('D1').value = 'StimDuration'
xlSheet.cell('E1').value = 'Block'



#####################
# Setup the Window. #
#####################

# Setup the Window
#win = visual.Window(size=(800, 768),
#                    fullscr=FULL_SCREEN,
#                    screen=0,
#                    allowGUI=False,
#                    allowStencil=False,
#                    monitor='monitor',
#                    color='black',
#                    colorSpace='rgb')

win = visual.Window(fullscr=FULL_SCREEN, monitor='testMonitor',color='black',)

instruct_text = visual.TextStim(win=win, ori=0, name='instruct_text',
    text=task_instructions,
    font='Arial',alignHoriz='center', alignVert='center',
    pos=[0, 0], height=0.1, wrapWidth=1.5,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)


fixation = visual.ShapeStim(win,units='cm', lineColor='white',
    lineWidth=1.2, vertices=((-0.9, 0), (0.9, 0), (0,0), (0,0.9), (0,-0.9)),
    interpolate=True, closeShape=False, pos=(0,0))

stimuli = visual.ImageStim(win=win, image = None)

thanks = visual.TextStim(win=win, ori=0, name='thanks',
    text='Thankyou',    font='Arial',
    pos=[0, 0], height=0.1, wrapWidth=1,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)


########################
# Start the experiment #
########################

globalClock = core.Clock()  # to track the time since experiment started
trialClock = core.Clock()
myMouse = event.Mouse()
win.setMouseVisible(False)
fixation.draw()
win.flip()
core.wait(0.5)
win.setMouseVisible(False)
instruct_text.draw()
win.flip()
core.wait(15)

for idx,pic in enumerate(range(1,10,1)):
    stimuli.setImage('test' + os.path.sep + 'PICTURE' +str(pic) + '.png')
    stimuli.draw()
    win.flip()
    core.wait(STIM_TIME)
    resp = get_response()
    if resp == 99:
        win.close()
        core.quit()
    win.flip()
    core.wait(ISI_TIME+random()*RND_TIME)


instruct_text.setText("Let's get started")
instruct_text.draw()
win.flip()
core.wait(5)
win.flip()


globalClock.reset()
pics = range(1,61,1)
shuffle(pics)
nEnsaio = 1
for idx,pic in enumerate(pics):
    stimuli.setImage('round1' + os.path.sep + 'PICTURE' +str(pic) + '.png')
    stimuli.draw()
    win.flip()
    trialClock.reset()
    onsetTime = globalClock.getTime()
    send_trigger(pic,parallel_port,parallel)
    while trialClock.getTime()<STIM_TIME-(.005):
        core.wait(.003)
        resp = get_response()
        if resp == 99:
            xlBook.save(filename + 'inc.xlsx')
            win.close()
            core.quit()
    #fixation.draw()
    win.flip()
    t_stim = trialClock.getTime()
    while trialClock.getTime()<(STIM_TIME +ISI_TIME + random()*RND_TIME)-(.005):
        core.wait(.003)
        resp = get_response()
        if resp == 99:
            xlBook.save(filename + 'inc.xlsx')
            win.close()
            core.quit()
    nEnsaio += 1
    xlSheet.cell('A'+str(nEnsaio)).value = idx+1
    xlSheet.cell('B'+str(nEnsaio)).value = 'PICTURE' +str(pic) + '.png'
    xlSheet.cell('C'+str(nEnsaio)).value = onsetTime
    xlSheet.cell('D'+str(nEnsaio)).value = t_stim
    xlSheet.cell('E'+str(nEnsaio)).value = 1

instruct_text.setText("Break\nPress to continue")
instruct_text.draw()
win.flip()
clear_all_cue()
while True:
    resp = get_response()
    if resp:
        if resp == 99:
            xlBook.save(filename + 'inc.xlsx')
            win.close()
            core.quit()
        else: break

win.flip()
core.wait(1)
win.flip()


pics = range(61,121,1)
shuffle(pics)

for idx,pic in enumerate(pics):
    stimuli.setImage('round2' + os.path.sep + 'PICTURE' +str(pic) + '.png')
    stimuli.draw()
    win.flip()
    trialClock.reset()
    onsetTime = globalClock.getTime()
    send_trigger(pic,parallel_port,parallel)
    while trialClock.getTime()<STIM_TIME-(.005):
        core.wait(.003)
        resp = get_response()
        if resp == 99:
            xlBook.save(filename + 'inc.xlsx')
            win.close()
            core.quit()
    #fixation.draw()
    win.flip()
    t_stim = trialClock.getTime()
    while trialClock.getTime()<(STIM_TIME +ISI_TIME+random()*RND_TIME)-(.005):
        core.wait(.003)
        resp = get_response()
        if resp == 99:
            xlBook.save(filename + 'inc.xlsx')
            win.close()
            core.quit()
    nEnsaio += 1
    xlSheet.cell('A'+str(nEnsaio+1)).value = idx+1
    xlSheet.cell('B'+str(nEnsaio+1)).value = 'PICTURE' +str(pic) + '.png'
    xlSheet.cell('C'+str(nEnsaio+1)).value = onsetTime
    xlSheet.cell('D'+str(nEnsaio+1)).value = t_stim
    xlSheet.cell('E'+str(nEnsaio+1)).value = 2




thanks.draw()
win.flip()
core.wait(2)
xlBook.save(filename + '.xlsx')
