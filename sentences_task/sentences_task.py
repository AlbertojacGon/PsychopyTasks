#!/usr/bin/env python
# -*- coding: utf-8 -*-


"""
Sentences task to run in Psychopy 1.85.4
Press esc to exit from the task
Relevant data is stored in an excel file inside "data" folder
"""

from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
from psychopy import visual, core, data, event, gui, parallel
import os, sys
from numpy.random import random, shuffle
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook




# Turn fullscreen off for testing on monitor that == not 1024x768
FULL_SCREEN = True

# CC here are some parameters that determine the behavior of the task
STIM_TIME = .35
ISI_TIME = 0.4
RND_ISI = 0.0 #random
BOX_TIME = 1.5
resp_text = 'GOOD OR BAD'

#Direccion do porto paralelo
try:
    parallel.setPortAddress(0x378)#direccion do porto paralelo
    parallel.setData(0)
    parallel_port=1
except:
    parallel_port=0
    print "Oops! No parallel port found"



def send_trigger(code,parallel_port,parallel):
    if parallel_port == 1:
        parallel.setData(code)
        core.wait(0.005)
        parallel.setData(0)
    else:
        pass


# clear all responses
def clear_all_cue():
    while True:
        if event.getKeys(): event.clearEvents('keyboard')
        else: break

# get responses
def get_response():
    key_out=[]
    resp=event.getKeys()
    if resp:
        key=resp[0]
        if key=="1":key_out=1
        elif key=="2":key_out=2
        elif key=="3":key_out=3
        elif key=="4":key_out=4
        elif key=="escape":key_out=99
        return key_out


task_instructions = u'In the first round, you will see 32 sentences. You have to decide whether the structure is semenatically correct or wrong. After you see "GOOD or BAD" on the screen, press the left key for "good" and the right key for "bad".'

expName = 'Senteces'  # from the Builder filename that created this script

# gui dialogue to get participant id, session number, and type of first
# block (useful for counterbalancing)
expInfo = {'1. ID Participant':'', '2. Group':''}
dlg = gui.DlgFromDict(dictionary=expInfo, title=expName, fixed=[])

# if user pressed cancel, quit
if dlg.OK == False:
    core.quit()

# set a few more configuration parameters
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName


# Setup files for saving
if not os.path.isdir('data'+ os.path.sep + '%s' %(expInfo['expName']) ):
    # if this fails (e.g. permissions) we will get error
    os.makedirs('data'+ os.path.sep + '%s' %(expInfo['expName']) )

filename = 'data' + os.path.sep + '%s' %(expInfo['expName']) \
    + os.path.sep + '%s_%s_%s' %(expInfo['1. ID Participant'],expInfo['2. Group'],\
                                    expInfo['date'])



#########
# EXCEL #
#########

xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]
xlSheet.title = 'picNaming' + str(expInfo['1. ID Participant'])

xlSheet.cell('A1').value = 'Trial'
xlSheet.cell('B1').value = 'Sentence'
xlSheet.cell('C1').value = 'Response'
xlSheet.cell('D1').value = 'Reaction Time'
xlSheet.cell('E1').value = 'Block'



#####################
# Setup the Window. #
#####################

win = visual.Window(fullscr=FULL_SCREEN, monitor='testMonitor',color='black',)

instruct_text = visual.TextStim(win=win, ori=0, name='instruct_text',
    text=task_instructions,
    font='Arial',alignHoriz='center', alignVert='center',
    pos=[0, 0], height=0.1, wrapWidth=1.5,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)


fixation = visual.ShapeStim(win,units='cm', lineColor='white',
    lineWidth=1.2, vertices=((-0.9, 0), (0.9, 0), (0,0), (0,0.9), (0,-0.9)),
    interpolate=True, closeShape=False, pos=(0,0))

box = visual.ShapeStim(win,units='cm', lineColor='white',
    lineWidth=1.2, vertices=((-1.9, -0.9), (1.9, -0.9), (1.9,0.9), (-1.9,0.9)),
    interpolate=True, closeShape=False, pos=(0,0))

response_text = visual.TextStim(win=win, ori=0, name='thanks',
    text= resp_text,    font='Arial',
    pos=[0, 0], height=0.1, wrapWidth=1,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)

word = visual.TextStim(win=win, ori=0, name='word',
    text='XXX',    font='Arial',
    pos=[0, 0], height=0.1, wrapWidth=1,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)

thanks = visual.TextStim(win=win, ori=0, name='thanks',
    text='Thankyou',    font='Arial',
    pos=[0, 0], height=0.1, wrapWidth=1,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)
win.setMouseVisible(False)



order1 = [0, 0, 1, 1, 0, 1, 1, 0, 1, 1, 0, 1, 0, 1, 1, 0, 0, 1, 1, 0, 0, 1, 0, 1, 0, 0, 0, 1, 1, 0, 1, 0, 0, 1, 1, 1, 1, 1,\
            0, 0, 0, 0, 1, 1, 0, 1, 1, 0, 0, 1, 0, 1, 0, 0, 0, 1, 1, 1, 0, 1, 0, 1, 1, 0, 0, 0, 0, 1, 1, 0,1, 0, 1, 0, 0, 0, 1, 1, 1, 0]

order2 = [1, 1, 0, 0, 1, 0, 0, 1, 0, 0, 1, 0, 1, 0, 0, 1, 1, 0, 0, 1, 1, 0, 1, 0, 1, 1, 1, 0, 0, 1, 0, 1, 1, 0, 0, 0, 0, 0,\
            1, 1, 1, 1, 0, 0, 1, 0, 0, 1, 1, 0, 1, 0, 1, 1, 1, 0, 0, 0, 1, 0, 1, 0, 0, 1, 1, 1, 1, 0, 0, 1, 0, 1, 0, 1, 1, 1, 0, 0, 0, 1]

try:
    if int(expInfo['2. Group'])%2 == 1:
        order = order1
        print 'order1'
    elif int(expInfo['2. Group'])%2 == 0:
        order = order2
        print 'order1'
except:
    print "The value '2. Group' has to be a number"
    win.close()
    core.quit()
########################
#      sentences       #
########################



wb = load_workbook(filename = 'sentences.xlsx')
ws = wb.active




########################
# Start the experiment #
########################

globalClock = core.Clock()  # to track the time since experiment started
trialClock = core.Clock()
respClock = core.Clock()
myMouse = event.Mouse()
win.setMouseVisible(False)
fixation.draw()
win.flip()
core.wait(0.5)
win.setMouseVisible(False)
instruct_text.draw()
win.flip()
core.wait(10)
win.flip()
globalClock.reset()
instruct_text.setText('Press a key to begin with some practice trials')
instruct_text.draw()
win.flip()
clear_all_cue()
while True:
        resp = get_response()
        if resp:
            break

for ridx, row in enumerate(ws.rows):
    sentence = []
    if ridx == 6: # if it is the sentence number 7
        win.flip()
        core.wait(1)
        instruct_text.setText("Let's get started")
        instruct_text.draw()
        win.flip()
        core.wait(5)
        win.flip()
    if ridx == 38: # if it is the sentence number 32
        win.flip()
        core.wait(1)
        instruct_text.setText("Break\nPress to continue")
        instruct_text.draw()
        win.flip()
        while True:
            resp = get_response()
            if resp:
                if resp == 99:
                    xlBook.save(filename + 'inc.xlsx')
                    win.close()
                    core.quit()
                else: break
        win.flip()
        core.wait(1)
    fixation.draw()
    win.flip()
    core.wait(0.5)
    clear_all_cue()

    cCont = 1
    for cidx, cell in enumerate(row):
        cCont += 1
        if cell.value:
            if cidx == 11:
                if order[ridx] == 0:
                    word.setText(cell.value)
                    sentence.append(cell.value)
                    word.draw()
                    win.flip()
                    trialClock.reset()
                    send_trigger(98,parallel_port,parallel)
                else: continue
            elif cidx == 12:
                if order[ridx] == 1:
                    word.setText(cell.value)
                    sentence.append(cell.value)
                    word.draw()
                    win.flip()
                    trialClock.reset()
                    send_trigger(99,parallel_port,parallel)
                else: continue
            else:
                word.setText(cell.value)
                sentence.append(cell.value)
                word.draw()
                win.flip()
                trialClock.reset()
                send_trigger(ridx+1,parallel_port,parallel)
            while trialClock.getTime()+.008 < STIM_TIME:
                core.wait(.002)
            win.flip()
            RTIME = random()*RND_ISI
            while trialClock.getTime() < (STIM_TIME+ISI_TIME+RTIME):
                core.wait(.002)

    respClock.reset()
    while respClock.getTime()<1.08:
        core.wait(.005)
    clear_all_cue()
    response_text.draw()
    win.flip()
    respClock.reset()
    send_trigger(100,parallel_port,parallel)
    rt = []
    response = []
    while respClock.getTime()<20.0:
        resp = get_response()
        if resp:
            rt = respClock.getTime()
            if resp == 99:
                xlBook.save(filename + 'inc.xlsx')
                win.close()
                core.quit()
            elif resp == 1:
                send_trigger(101,parallel_port,parallel)
                response = resp
            elif resp == 2:
                send_trigger(102,parallel_port,parallel)
                response = resp
            elif resp == 3:
                send_trigger(103,parallel_port,parallel)
                response = resp
            elif resp == 4:
                send_trigger(104,parallel_port,parallel)
                response = resp
            break
    sentence =[str(i) for i in sentence]
    xlSheet.cell('A'+str(ridx+2)).value = ridx+1
    xlSheet.cell('B'+str(ridx+2)).value = " ".join(sentence)
    if response:
        xlSheet.cell('C'+str(ridx+2)).value = response
        xlSheet.cell('D'+str(ridx+2)).value = rt
    win.flip()
    core.wait(0.5)

thanks.draw()
win.flip()
core.wait(2)
xlBook.save(filename + '.xlsx')
