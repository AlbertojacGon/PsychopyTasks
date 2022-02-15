 #!/usr/bin/env python2
# -*- coding: utf-8 -*-


from __future__ import division
import os, csv, random
from psychopy import visual, core, event, data, gui, parallel
from openpyxl import Workbook



# task instructions
instructionsBegin = u'Pressiona a tecla "enter" para começar'
instructionsBreak = u'Pausa. Pressiona a tecla "enter" para continuar.'
instructionsText1 = u'Quão agradável achas que foi a ação representada no vídeo anterior para a pessoa que estava a ser tocada?'
instructionsText2 = u'Quanto é que gostarias de ser tocado(a) da mesma forma?'

FULLSCR = True

expName = 'TouchVideos'

# stims (video, rating, trigger)
videos = [['Kate_hand_01.mp4', 1,201], ['Kate_hand_02.mp4', 1,202], ['Kate_hand_03.mp4',1,203],
         ['Kate_armA_01.mp4',1,211], ['Kate_armA_02.mp4',1,212], ['Kate_armA_03.mp4',1,213],
         ['Kate_armB_01.mp4',1,221], ['Kate_armB_02.mp4',1,222], ['Kate_armB_03.mp4',1,223],
         ['Kate_uparm_01.mp4',1,231], ['Kate_uparm_02.mp4',1,232], ['Kate_uparm_03.mp4',1,233],
         ['Kate_back_01.mp4',1,241], ['Kate_back_02.mp4',1,242], ['Kate_back_03.mp4',1,243],
         ['Kate_hand_01.mp4', 0,201], ['Kate_hand_02.mp4', 0,202], ['Kate_hand_03.mp4',0,203],
         ['Kate_armA_01.mp4',0,211], ['Kate_armA_02.mp4',0,212], ['Kate_armA_03.mp4',0,213],
         ['Kate_armB_01.mp4',0,221], ['Kate_armB_02.mp4',0,222], ['Kate_armB_03.mp4',0,223],
         ['Kate_uparm_01.mp4',0,231], ['Kate_uparm_02.mp4',0,232], ['Kate_uparm_03.mp4',0,233],
         ['Kate_back_01.mp4',0,241], ['Kate_back_02.mp4',0,242], ['Kate_back_03.mp4',0,243],
         ['Kate_hand_01.mp4', 0,201], ['Kate_hand_02.mp4', 0,202], ['Kate_hand_03.mp4',0,203],
         ['Kate_armA_01.mp4',0,211], ['Kate_armA_02.mp4',0,212], ['Kate_armA_03.mp4',0,213],
         ['Kate_armB_01.mp4',0,221], ['Kate_armB_02.mp4',0,222], ['Kate_armB_03.mp4',0,223],
         ['Kate_uparm_01.mp4',0,231], ['Kate_uparm_02.mp4',0,232], ['Kate_uparm_03.mp4',0,233],
         ['Kate_back_01.mp4',0,241], ['Kate_back_02.mp4',0,242], ['Kate_back_03.mp4',0,243],
         ['Kate_hand_01.mp4', 0,201], ['Kate_hand_02.mp4', 0,202], ['Kate_hand_03.mp4',0,203],
         ['Kate_armA_01.mp4',0,211], ['Kate_armA_02.mp4',0,212], ['Kate_armA_03.mp4',0,213],
         ['Kate_armB_01.mp4',0,221], ['Kate_armB_02.mp4',0,222], ['Kate_armB_03.mp4',0,223],
         ['Kate_uparm_01.mp4',0,231], ['Kate_uparm_02.mp4',0,232], ['Kate_uparm_03.mp4',0,233],
         ['Kate_back_01.mp4',0,241], ['Kate_back_02.mp4',0,242], ['Kate_back_03.mp4',0,243],
         ['Kate_hand_01.mp4', 0,201], ['Kate_hand_02.mp4', 0,202], ['Kate_hand_03.mp4',0,203],
         ['Kate_armA_01.mp4',0,211], ['Kate_armA_02.mp4',0,212], ['Kate_armA_03.mp4',0,213],
         ['Kate_armB_01.mp4',0,221], ['Kate_armB_02.mp4',0,222], ['Kate_armB_03.mp4',0,223],
         ['Kate_uparm_01.mp4',0,231], ['Kate_uparm_02.mp4',0,232], ['Kate_uparm_03.mp4',0,233],
         ['Kate_back_01.mp4',0,241], ['Kate_back_02.mp4',0,242], ['Kate_back_03.mp4',0,243]
         ]
random.shuffle(videos)


#parallel port setup
try:
    parallel.setPortAddress(0x378)#direccion do porto paralelo
    parallel.setData(0)
    parallel_port=1
except:
    parallel_port=0
    print("Oops!  No parallel port found")


# functions for responses
def clear_all_cue():
    while True:
        if event.getKeys(): event.clearEvents('keyboard')
        else: break

def get_response():
    key_out=[]
    resp=event.getKeys()
    if resp:
        key=resp[0]
        if key=='return':key_out=98
        elif key=="escape":key_out=99
        return key_out

def send_trigger(code,parallel_port,parallel):
    if parallel_port == 1:
        parallel.setData(code)
        core.wait(.005)
        parallel.setData(0)
    else:
        pass

#########
# EXCEL #
#########

def saveExcel(ntrials,movNames,rates1,rates2,rts1,rts2,totaltimes,videoDurs):
    dest_filename = 'data' + os.path.sep + '%s_%s_%s.xlsx' %(expInfo['1 SubjectNumber'], expInfo['2 Group'],expInfo['Date'])
    wb = Workbook()
    ws = wb.active
    ws.title = expName
    ws['A1'] = 'Trial'
    ws['B1'] = 'MovName'
    ws['C1'] = 'Rate1'
    ws['D1'] = 'Rate2'
    ws['E1'] = 'ReactionTime1'
    ws['F1'] = 'ReactionTime2'
    ws['G1'] = 'TotalTime'
    ws['H1'] = 'VideoDuration'
    for idx,ntrial in enumerate(ntrials):
        ws['A'+str(idx+2)] = ntrial
        ws['B'+str(idx+2)] = movNames[idx]
        ws['C'+str(idx+2)] = rates1[idx]
        ws['D'+str(idx+2)] = rates2[idx]
        ws['E'+str(idx+2)] = rts1[idx]
        ws['F'+str(idx+2)] = rts2[idx]
        ws['G'+str(idx+2)] = totaltimes[idx]
        ws['H'+str(idx+2)] = videoDurs[idx]
    wb.save(filename = dest_filename)



trialN = []
MovFile = []
rating1 = []
rating2 = []
RT1 = []
RT2 = []
totalT = []
VideoDurs = []


# initial gui
expInfo = {'1 SubjectNumber':'', '2 Group': ['Control', 'Test']}

infoDlg = gui.DlgFromDict(dictionary=expInfo, title= expName)
if not infoDlg.OK:
    print('User Cancelled')

# create a folder to save the data
expInfo['Date'] = data.getDateStr()
if not os.path.isdir('Data'):
    os.makedirs('Data')


win = visual.Window((800, 800), fullscr = FULLSCR,color = 'black')
fixation = visual.ShapeStim(win,lineWidth=6,vertices=((-0.05, 0), (0.05, 0), (0,0), (0,0.08), (0,-0.08)), interpolate = False, closeShape=False)
mov = visual.MovieStim3(win, 'Kate_uparm_02.mp4')
instruct = visual.TextStim(win=win, name='text',text= instructionsText1,font='Arial',
                          height=0.1, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, 0.5))
instruct2 = visual.TextStim(win=win, name='text',text= instructionsText1,font='Arial',
                          height=0.1, wrapWidth=1.8, color='white',alignHoriz='center',pos=(0.0, 0.0))

event.Mouse(visible = False)
mouse = event.Mouse(visible=False, newPos=None, win=win)
clear_all_cue()
instruct2.setText(instructionsBegin)
instruct2.draw()
win.flip()
while sum(mouse.getPressed())==0:
    core.wait(.1)


trialClock = core.Clock()
globalClock = core.Clock()


trialn = 0
for block in range(2):
    random.shuffle(videos)
    for video in videos:
        mov.setMovie(video[0])
        trialn += 1
        MovFile.append(video[0])
        trialN.append(trialn)
        totalT.append(globalClock.getTime())
        clear_all_cue()
        win.flip()
        send_trigger(video[2],parallel_port,parallel)
        T1=globalClock.getTime()
        while mov.status != visual.FINISHED:
            mov.draw()
            win.flip()
        T2=globalClock.getTime()
        fixation.draw()
        win.flip()
        core.wait(1+(random.random()*0.5))
        if video[1] == 1: # if a video that requires rating
            event.Mouse(visible = True)
            instruct.setText(instructionsText1)
            ratingScale1 = visual.RatingScale(win, low=1, high=7,precision=10,
                                            labels=['Muito desagradável', 'Muito agradável'], size=1.5,
                                            scale='1                                                7',
                                            acceptPreText='Aceitar')
            while ratingScale1.noResponse:
                resp = get_response()
                ratingScale1.draw()
                instruct.draw()
                win.flip()
            rating1temp = ratingScale1.getRating()
            send_trigger(int(rating1temp*10),parallel_port,parallel)
            instruct.setText(instructionsText2)
            ratingScale2 = visual.RatingScale(win, low=1, high=7,precision=10,
                                            labels=['De modo algum', 'Mesmo muito'], size=1.5,
                                            scale='1                                                7',
                                            acceptPreText='Aceitar')
            while ratingScale2.noResponse:
                resp = get_response()
                ratingScale2.draw()
                instruct.draw()
                win.flip()

            rating2temp = ratingScale2.getRating()
            send_trigger(int(rating2temp*10+100),parallel_port,parallel)
            rating1.append(rating1temp)
            RT1.append(ratingScale1.getRT())
            rating2.append(rating2temp)
            RT2.append(ratingScale2.getRT())
            VideoDurs.append(T2-T1)
        else:
            fixation.draw()
            win.flip()
            trialClock.reset()
            rating1.append('_')
            RT1.append('_')
            rating2.append('_')
            RT2.append('_')
            VideoDurs.append(T2-T1)
            while trialClock.getTime()<1.2:
                resp = get_response()
                if resp:
                    if resp == 99:
                        saveExcel(trialN,MovFile,rating1,rating2,RT1,RT2,totalT,VideoDurs)
                        win.close()
                        core.quit()
        event.Mouse(visible = False)
        while trialClock.getTime()<random.random()*0.5+0.7:
            resp = get_response()
            if resp:
                if resp == 99:
                    saveExcel(trialN,MovFile,rating1,rating2,RT1,RT2,totalT, VideoDurs )
                    win.close()
                    core.quit()

    if block == 0:
        event.clearEvents()
        instruct2.setText(instructionsBreak)
        instruct2.draw()
        win.flip()
        clear_all_cue()
        while sum(mouse.getPressed())==0:
            core.wait(.1)

instruct2.setText('Obrigado')
instruct2.draw()
win.flip()
saveExcel(trialN,MovFile,rating1,rating2,RT1,RT2,totalT,VideoDurs)
core.wait(2)
win.close()
core.quit()
